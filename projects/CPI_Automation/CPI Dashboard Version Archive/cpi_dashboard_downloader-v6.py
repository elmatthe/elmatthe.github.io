import tkinter as tk
from tkinter import messagebox, filedialog
import requests
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import time
import subprocess
import os
import openpyxl

# ========================================
# CONFIGURATION
# ========================================

FRED_APIKEY = "719fac656228747786ccc0d9f1a2ced2"
FRED_OBS_URL = "https://api.stlouisfed.org/fred/series/observations"

STATCAN_WDS_BASE = "https://www150.statcan.gc.ca/t1/wds/rest"

TARGET_SHEET = "CPI_Downloader"

# Canada series
STATCAN_VECTORS = {
    "Headline CPI - Unadjusted CDN": "41690973",
    "Headline CPI - Seasonally Adjusted CDN": "41690914",
    "Core CPI - Unadjusted CDN": "41691233",
    "Core CPI - Seasonally Adjusted CDN": "41690924",
    "Alberta CPI - Unadjusted CDN": "41692327",
}
STATCAN_COLUMNS = list(STATCAN_VECTORS.keys())

# US monthly series
FRED_MONTHLY_SERIES = {
    "Headline CPI - Unadjusted US": "CPIAUCNS",
    "Headline CPI - Seasonally Adjusted US": "CPIAUCSL",
    "Core CPI - Unadjusted US": "CPILFENS",
    "Core CPI - Seasonally Adjusted US": "CPILFESL",
    "Unemployment Rate - Seasonally Adjusted US": "UNRATE",
    "Monthly Effective Fed Fund Rate US": "FEDFUNDS",
    "PCE - Seasonally Adjusted US": "PCEPI",
    "Core PCE - Seasonally Adjusted US": "PCEPILFE",
}

# US daily end-of-month series
FRED_DAILY_EOM_SERIES = {
    "Monthly Fed Fund Rate - EOP US": "DFEDTARU",
    "SP by month US": "SP500",
    "10yr vs 2yr US": "T10Y2Y",
}

# ========================================
# REQUIRED COLUMN ORDER (A–Q)
# ========================================
OUTPUT_COLUMNS = [
    "Date",
    "Headline CPI - Unadjusted CDN",
    "Headline CPI - Seasonally Adjusted CDN",
    "Core CPI - Unadjusted CDN",
    "Core CPI - Seasonally Adjusted CDN",
    "Alberta CPI - Unadjusted CDN",
    "Headline CPI - Unadjusted US",
    "Headline CPI - Seasonally Adjusted US",
    "Core CPI - Unadjusted US",
    "Core CPI - Seasonally Adjusted US",
    "Unemployment Rate - Seasonally Adjusted US",
    "Monthly Effective Fed Fund Rate US",
    "Monthly Fed Fund Rate - EOP US",
    "SP by month US",
    "10yr vs 2yr US",
    "PCE - Seasonally Adjusted US",
    "Core PCE - Seasonally Adjusted US",
]

# ========================================
# ONEDRIVE PAUSE / RESUME
# ========================================

def _get_onedrive_exe() -> str | None:
    """
    Return the path to OneDrive.exe if it exists on this machine,
    otherwise return None (so we silently skip pause/resume on machines
    that don't have OneDrive installed).
    """
    candidate = os.path.join(
        os.environ.get("LOCALAPPDATA", ""),
        "Microsoft", "OneDrive", "OneDrive.exe"
    )
    return candidate if os.path.isfile(candidate) else None


def onedrive_pause():
    """
    Pause OneDrive syncing. Safe to call even if OneDrive is not installed —
    it will simply do nothing.
    """
    exe = _get_onedrive_exe()
    if exe:
        try:
            subprocess.run([exe, "/pause"], check=False,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            # Give OneDrive a moment to actually pause before we touch the file
            time.sleep(2)
        except Exception:
            pass  # Never let a OneDrive issue block the main task


def onedrive_resume():
    """
    Resume OneDrive syncing. Safe to call even if OneDrive is not installed.
    """
    exe = _get_onedrive_exe()
    if exe:
        try:
            subprocess.run([exe, "/resume"], check=False,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

# ========================================
# DATE HELPERS
# ========================================

def ym_to_date(ym: str) -> date:
    """Convert 'YYYY-MM' to a Python date object (first of month)."""
    year, month = ym.split("-")
    return date(int(year), int(month), 1)

def month_range(start_ym: str, end_ym: str):
    start = datetime.strptime(start_ym, "%Y-%m")
    end = datetime.strptime(end_ym, "%Y-%m")
    cur = start
    while cur <= end:
        yield cur.strftime("%Y-%m")
        cur += relativedelta(months=1)

# ========================================
# STATCAN FETCH
# ========================================

def fetch_statcan_vector_range(vector_id: int, start_ym: str, end_ym: str, max_retries: int = 3):
    start_date = f"{start_ym}-01"
    end_date = f"{end_ym}-01"

    url = f"{STATCAN_WDS_BASE}/getDataFromVectorByReferencePeriodRange"
    params = {
        "vectorIds": str(vector_id),
        "startRefPeriod": start_date,
        "endReferencePeriod": end_date,
    }

    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, timeout=120)
            r.raise_for_status()
            j = r.json()

            data = {}
            if isinstance(j, list) and j and j[0].get("status") == "SUCCESS":
                obj = j[0].get("object", {})
                for dp in obj.get("vectorDataPoint", []):
                    ref = dp.get("refPer")
                    val = dp.get("value")
                    if not ref or val is None:
                        continue
                    ym = ref[:7]
                    if ym < start_ym or ym > end_ym:
                        continue
                    try:
                        data[ym] = float(val)
                    except ValueError:
                        continue
            return data

        except requests.exceptions.RequestException:
            if attempt == max_retries - 1:
                raise
            time.sleep(2 ** attempt)

def fetch_all_statcan(start_ym: str, end_ym: str):
    out = {}
    for col, vec_id in STATCAN_VECTORS.items():
        out[col] = fetch_statcan_vector_range(int(vec_id), start_ym, end_ym)
    return out

# ========================================
# FRED FETCH
# ========================================

def fetch_fred_monthly(series_id: str, start_ym: str, end_ym: str, max_retries: int = 3):
    params = {
        "series_id": series_id,
        "api_key": FRED_APIKEY,
        "file_type": "json",
        "frequency": "m",
        "aggregation_method": "avg",
        "observation_start": f"{start_ym}-01",
        "observation_end": f"{end_ym}-28",
    }

    for attempt in range(max_retries):
        try:
            r = requests.get(FRED_OBS_URL, params=params, timeout=120)
            r.raise_for_status()
            j = r.json()
            out = {}
            for obs in j.get("observations", []):
                dt = obs["date"]
                ym = dt[:7]
                val_str = obs.get("value")
                if not val_str or val_str == ".":
                    continue
                try:
                    out[ym] = float(val_str)
                except ValueError:
                    continue
            return out

        except requests.exceptions.RequestException:
            if attempt == max_retries - 1:
                raise
            time.sleep(2 ** attempt)

def last_day_of_month(d: date) -> date:
    return d.replace(day=1) + relativedelta(months=1) + relativedelta(days=-1)

def fetch_fred_daily_eom(series_id: str, start_ym: str, end_ym: str, max_retries: int = 3):
    start_date = datetime.strptime(start_ym + "-01", "%Y-%m-%d").date()
    end_date = last_day_of_month(datetime.strptime(end_ym + "-01", "%Y-%m-%d").date())

    params = {
        "series_id": series_id,
        "api_key": FRED_APIKEY,
        "file_type": "json",
        "frequency": "d",
        "observation_start": start_date.isoformat(),
        "observation_end": end_date.isoformat(),
    }

    for attempt in range(max_retries):
        try:
            r = requests.get(FRED_OBS_URL, params=params, timeout=120)
            r.raise_for_status()
            j = r.json()
            tmp = {}
            for obs in j.get("observations", []):
                dt_str = obs["date"]
                val_str = obs.get("value")
                if not val_str or val_str == ".":
                    continue
                try:
                    val = float(val_str)
                except ValueError:
                    continue
                ym = dt_str[:7]
                d_val = datetime.strptime(dt_str, "%Y-%m-%d").date()
                if ym not in tmp or d_val > tmp[ym][0]:
                    tmp[ym] = (d_val, val)
            out = {ym: v for ym, (d_val, v) in tmp.items()}
            return out

        except requests.exceptions.RequestException:
            if attempt == max_retries - 1:
                raise
            time.sleep(2 ** attempt)

# ========================================
# BUILD ROWS
# ========================================

def build_rows(start_ym: str, end_ym: str):
    statcan_data = fetch_all_statcan(start_ym, end_ym)
    fred_monthly = {
        name: fetch_fred_monthly(sid, start_ym, end_ym)
        for name, sid in FRED_MONTHLY_SERIES.items()
    }
    fred_eom = {
        name: fetch_fred_daily_eom(sid, start_ym, end_ym)
        for name, sid in FRED_DAILY_EOM_SERIES.items()
    }

    rows = []
    for ym in month_range(start_ym, end_ym):
        row = {col: None for col in OUTPUT_COLUMNS}

        # Date as a real Python date object — Excel stores it as a date serial
        row["Date"] = ym_to_date(ym)

        for col in STATCAN_COLUMNS:
            if ym in statcan_data.get(col, {}):
                row[col] = statcan_data[col][ym]

        for col_name, series in fred_monthly.items():
            if ym in series:
                row[col_name] = series[ym]

        for col_name, series in fred_eom.items():
            if ym in series:
                row[col_name] = series[ym]

        rows.append(row)

    return rows

# ========================================
# WRITE DIRECTLY TO XLSM
# ========================================

def write_to_xlsm(rows, xlsm_path: str):
    """
    Pause OneDrive, open the workbook (preserving VBA), clear the
    CPI_Downloader sheet, write headers + data, save, then resume OneDrive.

    Resuming happens in a finally block so OneDrive is always re-enabled
    even if an error occurs mid-write.
    """
    onedrive_pause()

    try:
        wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)

        if TARGET_SHEET not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{TARGET_SHEET}' not found in workbook.\n"
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[TARGET_SHEET]

        # Clear all existing content
        ws.delete_rows(1, ws.max_row)

        # Write header row
        ws.append(OUTPUT_COLUMNS)

        # Write data rows in column order
        for row in rows:
            ws.append([row[col] for col in OUTPUT_COLUMNS])

        wb.save(xlsm_path)

    finally:
        # Always resume OneDrive — even if the save failed
        onedrive_resume()

# ========================================
# GUI
# ========================================

def run_download():
    mode = var_mode.get()
    y = entry_year.get().strip()
    m = entry_month.get().strip()
    y2 = entry_year2.get().strip()
    m2 = entry_month2.get().strip()
    xlsm_path = entry_xlsm.get().strip()

    try:
        # Validate date inputs
        if mode == "single":
            if not (y and m):
                raise ValueError("Enter year and month.")
            start_ym = end_ym = f"{int(y):04d}-{int(m):02d}"
        else:
            if not (y and m and y2 and m2):
                raise ValueError("Enter both start and end year/month.")
            start_ym = f"{int(y):04d}-{int(m):02d}"
            end_ym = f"{int(y2):04d}-{int(m2):02d}"
            if end_ym < start_ym:
                raise ValueError("End date must be after start date.")

        # Validate workbook path
        if not xlsm_path:
            raise ValueError("Select the target .xlsm workbook.")
        if not xlsm_path.lower().endswith((".xlsm", ".xlsx")):
            raise ValueError("Selected file must be a .xlsm or .xlsx workbook.")

        btn.config(state="disabled", text="Downloading…")
        root.update_idletasks()

        rows = build_rows(start_ym, end_ym)
        write_to_xlsm(rows, xlsm_path)

        messagebox.showinfo(
            "Done",
            f"✅  {len(rows)} rows written to the '{TARGET_SHEET}' sheet.\n\n"
            f"File: {xlsm_path}"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))

    finally:
        btn.config(state="normal", text="Download & Write to Excel")

def browse_xlsm():
    path = filedialog.askopenfilename(
        title="Select your .xlsm workbook",
        filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"),
                   ("Excel Workbook", "*.xlsx"),
                   ("All files", "*.*")]
    )
    if path:
        entry_xlsm.delete(0, tk.END)
        entry_xlsm.insert(0, path)

def on_mode_change():
    if var_mode.get() == "single":
        entry_year2.config(state="disabled")
        entry_month2.config(state="disabled")
    else:
        entry_year2.config(state="normal")
        entry_month2.config(state="normal")

def main():
    global entry_year, entry_month, entry_year2, entry_month2, var_mode
    global entry_xlsm, btn, root

    root = tk.Tk()
    root.title("CPI Dashboard Downloader — Direct to Excel")
    root.resizable(False, False)

    frame = tk.Frame(root, padx=14, pady=14)
    frame.pack()

    # ── Mode ──────────────────────────────────────────────────────────────────
    var_mode = tk.StringVar(value="single")
    tk.Label(frame, text="Mode:").grid(row=0, column=0, sticky="w")
    tk.Radiobutton(frame, text="Single month", variable=var_mode,
                   value="single", command=on_mode_change).grid(row=0, column=1, sticky="w")
    tk.Radiobutton(frame, text="Date range", variable=var_mode,
                   value="range", command=on_mode_change).grid(row=0, column=2, sticky="w")

    # ── Start date ────────────────────────────────────────────────────────────
    tk.Label(frame, text="Start Year (YYYY):").grid(row=1, column=0, sticky="w", pady=(8, 0))
    entry_year = tk.Entry(frame, width=6)
    entry_year.grid(row=1, column=1, sticky="w", pady=(8, 0))

    tk.Label(frame, text="Start Month (1-12):").grid(row=1, column=2, sticky="w", pady=(8, 0))
    entry_month = tk.Entry(frame, width=4)
    entry_month.grid(row=1, column=3, sticky="w", pady=(8, 0))

    # ── End date ──────────────────────────────────────────────────────────────
    tk.Label(frame, text="End Year (YYYY):").grid(row=2, column=0, sticky="w")
    entry_year2 = tk.Entry(frame, width=6)
    entry_year2.grid(row=2, column=1, sticky="w")

    tk.Label(frame, text="End Month (1-12):").grid(row=2, column=2, sticky="w")
    entry_month2 = tk.Entry(frame, width=4)
    entry_month2.grid(row=2, column=3, sticky="w")

    on_mode_change()

    # ── Workbook picker ───────────────────────────────────────────────────────
    tk.Label(frame, text="Target .xlsm:").grid(row=3, column=0, sticky="w", pady=(12, 0))
    entry_xlsm = tk.Entry(frame, width=42)
    entry_xlsm.grid(row=3, column=1, columnspan=2, sticky="ew", pady=(12, 0))
    tk.Button(frame, text="Browse…", command=browse_xlsm).grid(
        row=3, column=3, sticky="w", pady=(12, 0), padx=(4, 0))

    # ── Action button ─────────────────────────────────────────────────────────
    btn = tk.Button(frame, text="Download & Write to Excel",
                    command=run_download, bg="#1a6faf", fg="white",
                    font=("Segoe UI", 10, "bold"), padx=8, pady=4)
    btn.grid(row=4, column=0, columnspan=4, pady=(16, 0))

    root.mainloop()

if __name__ == "__main__":
    main()