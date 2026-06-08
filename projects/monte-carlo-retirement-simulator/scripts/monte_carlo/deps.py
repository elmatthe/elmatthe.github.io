"""Optional third-party dependencies, isolated in one place.

The engine only needs numpy; the Excel/CSV report writer also needs openpyxl and
matplotlib. Each import is guarded so the program can give a clear, friendly message
instead of crashing with an ImportError when something is missing.
"""
from __future__ import annotations

MISSING_DEPENDENCIES: list[str] = []

try:
    import numpy as np
except ImportError:
    np = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("numpy")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]
    XLImage = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    MISSING_DEPENDENCIES.append("openpyxl")

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter
except ImportError:
    plt = None  # type: ignore[assignment]
    FuncFormatter = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("matplotlib")


NUMPY_AVAILABLE = np is not None
EXCEL_AVAILABLE = Workbook is not None
CHART_AVAILABLE = plt is not None


def require_numpy() -> None:
    """Raise if numpy (the only dependency the simulation engine needs) is missing."""
    if not NUMPY_AVAILABLE:
        raise RuntimeError(
            "Missing required dependency: numpy. "
            "Install it with: pip install numpy"
        )


def require_export_deps() -> None:
    """Raise if the Excel/chart export dependencies are missing."""
    missing = [name for name, ok in (("openpyxl", EXCEL_AVAILABLE), ("matplotlib", CHART_AVAILABLE)) if not ok]
    if missing:
        raise RuntimeError(
            f"Missing required dependencies for report export: {', '.join(missing)}. "
            "Install them with: pip install openpyxl matplotlib"
        )


def require_all_dependencies() -> None:
    """Raise if any dependency is missing (used by the desktop run flow)."""
    if MISSING_DEPENDENCIES:
        deps = ", ".join(MISSING_DEPENDENCIES)
        raise RuntimeError(
            f"Missing required dependencies: {deps}. "
            "Install them with: pip install openpyxl matplotlib numpy"
        )
