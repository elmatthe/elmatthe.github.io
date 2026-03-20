#!/usr/bin/env python3
"""
Portfolio Rebalancer Desktop Tool
---------------------------------
Desktop tool for portfolio rebalancing with:
- Multi-currency FX conversion
- Optional live ticker/FX fetch (Yahoo Finance)
- Optional CSV/Excel export
"""

from __future__ import annotations

import csv
import datetime as dt
import queue
import threading
from pathlib import Path

TK_AVAILABLE = True
try:
    import tkinter as tk
    from tkinter import filedialog, ttk
except ImportError:
    TK_AVAILABLE = False
    tk = None  # type: ignore[assignment]
    filedialog = None  # type: ignore[assignment]
    ttk = None  # type: ignore[assignment]

MISSING_DEPENDENCIES: list[str] = []

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("openpyxl")

try:
    import yfinance as yf
except ImportError:
    yf = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("yfinance")


STATUS_NEUTRAL = "#1f4f8a"
STATUS_ERROR = "#9b2f2f"
STATUS_SUCCESS = "#1c7550"
STATUS_WARN = "#8f5f12"
BRAND_NAVY_ARGB = "FF1A2E4A"
HOLD_THRESHOLD = 0.005

EXCHANGE_SUFFIX_HINTS = {
    "USD": ["", ".US"],
    "CAD": [".TO", ".V", ""],
    "JPN": [".T", ""],
    "EUR": [".DE", ".AS", ".PA", ".MI", ".BR", ""],
    "GBP": [".L", ""],
    "CHY_CNH": [".SS", ".SZ", ".HK", ""],
}

CURRENCY_OPTIONS = {
    "USD": {"code": "USD", "label": "USD", "fx_to_usd": 1.00, "symbol": "$"},
    "CAD": {"code": "CAD", "label": "CAD", "fx_to_usd": 0.74, "symbol": "C$"},
    "JPN": {"code": "JPY", "label": "JPN", "fx_to_usd": 0.0067, "symbol": "JPY "},
    "EUR": {"code": "EUR", "label": "EUR", "fx_to_usd": 1.09, "symbol": "EUR "},
    "GBP": {"code": "GBP", "label": "GBP", "fx_to_usd": 1.28, "symbol": "GBP "},
    "CHY_CNH": {"code": "CNY", "label": "CHY/CNH", "fx_to_usd": 0.14, "symbol": "CNY "},
}

CURRENCY_CODE_TO_KEY = {meta["code"]: key for key, meta in CURRENCY_OPTIONS.items()}

SAMPLE_PORTFOLIO = [
    {"ticker": "VTI", "shares": 120, "price": 250, "targetWeight": 10, "currencyKey": "USD"},
    {"ticker": "XIC", "shares": 320, "price": 36, "targetWeight": 10, "currencyKey": "CAD"},
    {"ticker": "VEQT", "shares": 410, "price": 42, "targetWeight": 20, "currencyKey": "CAD"},
    {"ticker": "EWJ", "shares": 220, "price": 2480, "targetWeight": 10, "currencyKey": "JPN"},
    {"ticker": "VGK", "shares": 115, "price": 64, "targetWeight": 10, "currencyKey": "EUR"},
    {"ticker": "ISF", "shares": 260, "price": 7.4, "targetWeight": 10, "currencyKey": "GBP"},
    {"ticker": "AAPL", "shares": 45, "price": 180, "targetWeight": 10, "currencyKey": "USD"},
    {"ticker": "BND", "shares": 200, "price": 72, "targetWeight": 10, "currencyKey": "USD"},
    {"ticker": "MCHI", "shares": 140, "price": 40, "targetWeight": 10, "currencyKey": "USD"},
]


def clamp_row_count(value: int) -> int:
    return max(1, min(50, int(value)))


def _replace_or_create_sheet(workbook: "Workbook", title: str):
    if title in workbook.sheetnames:
        idx = workbook.sheetnames.index(title)
        del workbook[title]
        return workbook.create_sheet(title=title, index=idx)
    return workbook.create_sheet(title=title)


def _safe_positive_float(raw_value: object) -> float | None:
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    return value


def _extract_yf_last_price(yf_ticker) -> float | None:
    fast_info = getattr(yf_ticker, "fast_info", None)
    if fast_info is not None:
        for key in ("last_price", "regular_market_price", "previous_close"):
            raw = None
            try:
                if hasattr(fast_info, "get"):
                    raw = fast_info.get(key)
                else:
                    raw = getattr(fast_info, key, None)
            except Exception:
                raw = None
            value = _safe_positive_float(raw)
            if value is not None:
                return value

    try:
        history = yf_ticker.history(period="5d", interval="1d")
        if not history.empty:
            raw = history["Close"].dropna().iloc[-1]
            value = _safe_positive_float(raw)
            if value is not None:
                return value
    except Exception:
        pass

    return None


def _extract_yf_quote_currency(yf_ticker) -> str | None:
    fast_info = getattr(yf_ticker, "fast_info", None)
    if fast_info is not None:
        for key in ("currency", "currency_code"):
            raw_value = None
            try:
                if hasattr(fast_info, "get"):
                    raw_value = fast_info.get(key)
                else:
                    raw_value = getattr(fast_info, key, None)
            except Exception:
                raw_value = None
            if isinstance(raw_value, str) and raw_value.strip():
                return raw_value.strip()

    try:
        info = yf_ticker.get_info()
        if isinstance(info, dict):
            raw_value = info.get("currency")
            if isinstance(raw_value, str) and raw_value.strip():
                return raw_value.strip()
    except Exception:
        pass

    return None


def _normalize_quote_currency(raw_currency: str | None) -> tuple[str | None, float]:
    if raw_currency is None:
        return None, 1.0
    cleaned = raw_currency.strip()
    if not cleaned:
        return None, 1.0

    # LSE quotes often use pence labels (GBp/GBX).
    if cleaned in {"GBp", "GBX"}:
        return "GBP", 0.01

    upper = cleaned.upper()
    if upper == "CNH":
        return "CNY", 1.0
    return upper, 1.0


def _build_ticker_candidates(ticker_symbol: str, currency_key: str) -> list[str]:
    base = ticker_symbol.strip().upper()
    if not base:
        return []
    if "." in base or "=" in base:
        return [base]

    candidates: list[str] = []
    for suffix in EXCHANGE_SUFFIX_HINTS.get(currency_key, [""]):
        candidate = f"{base}{suffix}" if suffix else base
        if candidate not in candidates:
            candidates.append(candidate)
    return candidates


def fetch_live_quote_details(ticker_symbol: str) -> dict | None:
    if yf is None:
        return None
    try:
        ticker = yf.Ticker(ticker_symbol)
        raw_price = _extract_yf_last_price(ticker)
        if raw_price is None:
            return None

        raw_quote_currency = _extract_yf_quote_currency(ticker)
        quote_currency, unit_scale = _normalize_quote_currency(raw_quote_currency)
        return {
            "resolvedTicker": ticker_symbol,
            "rawPrice": raw_price,
            "price": raw_price * unit_scale,
            "rawQuoteCurrency": raw_quote_currency,
            "quoteCurrency": quote_currency,
            "unitScale": unit_scale,
        }
    except Exception:
        return None


def fetch_live_price(ticker_symbol: str) -> float | None:
    if yf is None:
        return None
    try:
        ticker = yf.Ticker(ticker_symbol)
        return _extract_yf_last_price(ticker)
    except Exception:
        return None


def fetch_live_fx_to_usd(currency_code: str) -> float | None:
    if currency_code == "USD":
        return 1.0
    if yf is None:
        return None
    pair_symbol = f"{currency_code}USD=X"
    try:
        ticker = yf.Ticker(pair_symbol)
        return _extract_yf_last_price(ticker)
    except Exception:
        return None


def calculate_rebalance_plan(positions: list[dict], net_flow: float) -> tuple[dict, list[dict]]:
    total_current = sum(pos["currentValue"] for pos in positions)
    total_weight = sum(pos["targetWeight"] for pos in positions)
    ending_value = total_current + net_flow
    if ending_value <= 0:
        raise ValueError("Ending portfolio value must be greater than 0.")
    if total_weight <= 0:
        raise ValueError("At least one target weight must be greater than 0.")

    total_buys = 0.0
    total_sells = 0.0
    results = []

    for pos in positions:
        target_weight_norm = pos["targetWeight"] / total_weight
        target_value = target_weight_norm * ending_value
        trade_value = target_value - pos["currentValue"]
        trade_value_local = trade_value / pos["fxToReporting"]
        trade_shares = trade_value_local / pos["price"]

        action = "Hold"
        if trade_value > HOLD_THRESHOLD:
            action = "Buy"
            total_buys += trade_value
        elif trade_value < -HOLD_THRESHOLD:
            action = "Sell"
            total_sells += abs(trade_value)

        if abs(trade_value) <= HOLD_THRESHOLD:
            trade_value = 0.0
            trade_value_local = 0.0
            trade_shares = 0.0

        results.append(
            {
                "ticker": pos["ticker"],
                "currencyKey": pos["currencyKey"],
                "currencyLabel": pos["currencyLabel"],
                "shares": pos["shares"],
                "price": pos["price"],
                "currentValue": pos["currentValue"],
                "targetWeightNorm": target_weight_norm,
                "targetValue": target_value,
                "tradeValue": trade_value,
                "tradeValueLocal": trade_value_local,
                "tradeShares": trade_shares,
                "action": action,
                "postTradeShares": pos["shares"] + trade_shares,
            }
        )

    summary = {
        "runDate": dt.datetime.now().isoformat(timespec="seconds"),
        "totalCurrent": total_current,
        "netFlow": net_flow,
        "endingValue": ending_value,
        "totalBuys": total_buys,
        "totalSells": total_sells,
    }
    return summary, results


def write_csv_export(csv_path: Path, summary: dict, rows: list[dict], reporting_currency: str) -> None:
    with csv_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["Run Date", summary["runDate"]])
        writer.writerow(["Reporting Currency", reporting_currency])
        writer.writerow(["Total Current Value", f"{summary['totalCurrent']:.2f}"])
        writer.writerow(["Net Flow", f"{summary['netFlow']:.2f}"])
        writer.writerow(["Target Ending Value", f"{summary['endingValue']:.2f}"])
        writer.writerow(["Total Buy Value", f"{summary['totalBuys']:.2f}"])
        writer.writerow(["Total Sell Value", f"{summary['totalSells']:.2f}"])
        writer.writerow([])
        writer.writerow(
            [
                "Ticker",
                "Currency",
                "Shares",
                "Price (Local)",
                "Current Value",
                "Target %",
                "Target Value",
                "Trade (Reporting)",
                "Trade (Local)",
                "Trade Shares",
                "Action",
                "Post-Trade Shares",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row["ticker"],
                    row["currencyLabel"],
                    f"{row['shares']:.4f}",
                    f"{row['price']:.4f}",
                    f"{row['currentValue']:.2f}",
                    f"{row['targetWeightNorm'] * 100:.2f}",
                    f"{row['targetValue']:.2f}",
                    f"{row['tradeValue']:.2f}",
                    f"{row['tradeValueLocal']:.2f}",
                    f"{row['tradeShares']:.4f}",
                    row["action"],
                    f"{row['postTradeShares']:.4f}",
                ]
            )


def write_excel_export(excel_path: Path, summary: dict, rows: list[dict], reporting_currency: str) -> None:
    if Workbook is None or load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")
    assert Font is not None and PatternFill is not None and Alignment is not None

    workbook_is_new = not excel_path.exists()
    workbook = load_workbook(excel_path) if excel_path.exists() else Workbook()
    summary_sheet = _replace_or_create_sheet(workbook, "RB_Summary")
    trades_sheet = _replace_or_create_sheet(workbook, "RB_TradePlan")

    if workbook_is_new and "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    navy_fill = PatternFill(fill_type="solid", start_color=BRAND_NAVY_ARGB, end_color=BRAND_NAVY_ARGB)
    white_bold = Font(color="FFFFFF", bold=True)
    heading_font = Font(bold=True)

    summary_sheet.merge_cells("A1:B1")
    header = summary_sheet["A1"]
    header.value = "Portfolio Rebalancer Summary"
    header.fill = navy_fill
    header.font = white_bold
    header.alignment = Alignment(horizontal="center")

    summary_rows = [
        (3, "Run Date", summary["runDate"], None),
        (4, "Reporting Currency", reporting_currency, None),
        (5, "Total Current Value", summary["totalCurrent"], "$#,##0.00"),
        (6, "Net Flow", summary["netFlow"], "$#,##0.00"),
        (7, "Target Ending Value", summary["endingValue"], "$#,##0.00"),
        (8, "Total Buy Value", summary["totalBuys"], "$#,##0.00"),
        (9, "Total Sell Value", summary["totalSells"], "$#,##0.00"),
    ]

    for row_idx, label, value, number_format in summary_rows:
        summary_sheet.cell(row=row_idx, column=1, value=label).font = heading_font
        value_cell = summary_sheet.cell(row=row_idx, column=2, value=value)
        if number_format:
            value_cell.number_format = number_format

    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 26

    headers = [
        "Ticker",
        "Currency",
        "Shares",
        "Price (Local)",
        "Current Value",
        "Target %",
        "Target Value",
        "Trade (Reporting)",
        "Trade (Local)",
        "Trade Shares",
        "Action",
        "Post-Trade Shares",
    ]

    header_fill = PatternFill(fill_type="solid", start_color=BRAND_NAVY_ARGB, end_color=BRAND_NAVY_ARGB)
    for col_idx, label in enumerate(headers, start=1):
        cell = trades_sheet.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = white_bold

    for row_idx, row in enumerate(rows, start=2):
        trades_sheet.cell(row=row_idx, column=1, value=row["ticker"])
        trades_sheet.cell(row=row_idx, column=2, value=row["currencyLabel"])
        trades_sheet.cell(row=row_idx, column=3, value=float(row["shares"])).number_format = "#,##0.0000"
        trades_sheet.cell(row=row_idx, column=4, value=float(row["price"])).number_format = "#,##0.0000"
        trades_sheet.cell(row=row_idx, column=5, value=float(row["currentValue"])).number_format = "$#,##0.00"
        trades_sheet.cell(row=row_idx, column=6, value=float(row["targetWeightNorm"])).number_format = "0.00%"
        trades_sheet.cell(row=row_idx, column=7, value=float(row["targetValue"])).number_format = "$#,##0.00"
        trades_sheet.cell(row=row_idx, column=8, value=float(row["tradeValue"])).number_format = "$#,##0.00"
        trades_sheet.cell(row=row_idx, column=9, value=float(row["tradeValueLocal"])).number_format = "#,##0.00"
        trades_sheet.cell(row=row_idx, column=10, value=float(row["tradeShares"])).number_format = "#,##0.0000"
        action_cell = trades_sheet.cell(row=row_idx, column=11, value=row["action"])
        trades_sheet.cell(row=row_idx, column=12, value=float(row["postTradeShares"])).number_format = "#,##0.0000"

        if row["action"] == "Buy":
            action_cell.fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        elif row["action"] == "Sell":
            action_cell.fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    trades_sheet.column_dimensions["A"].width = 12
    trades_sheet.column_dimensions["B"].width = 12
    trades_sheet.column_dimensions["C"].width = 14
    trades_sheet.column_dimensions["D"].width = 14
    trades_sheet.column_dimensions["E"].width = 14
    trades_sheet.column_dimensions["F"].width = 12
    trades_sheet.column_dimensions["G"].width = 14
    trades_sheet.column_dimensions["H"].width = 16
    trades_sheet.column_dimensions["I"].width = 14
    trades_sheet.column_dimensions["J"].width = 14
    trades_sheet.column_dimensions["K"].width = 10
    trades_sheet.column_dimensions["L"].width = 16
    trades_sheet.freeze_panes = "A2"

    try:
        workbook.save(excel_path)
    except PermissionError as exc:
        raise PermissionError("The workbook appears to be open. Close it in Excel and try again.") from exc


class PortfolioRebalancerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Portfolio Rebalancer Desktop Tool")
        self.root.geometry("1360x920")
        self.root.minsize(1180, 760)

        self.rows = []
        self.has_calculated = False
        self.is_running = False
        self.worker_thread: threading.Thread | None = None
        self.worker_events: queue.Queue[tuple[str, object]] = queue.Queue()

        self.row_count_var = tk.StringVar(value=str(len(SAMPLE_PORTFOLIO)))
        self.reporting_currency_var = tk.StringVar(value="USD")
        self.net_flow_var = tk.StringVar(value="0")
        self.validation_var = tk.StringVar(value="")
        self.live_current_total_var = tk.StringVar(value="Current total: $0.00")
        self.live_weight_total_var = tk.StringVar(value="Target weight total: 0.00%")
        self.net_flow_label_var = tk.StringVar(value="Net contribution / withdrawal (USD)")
        self.fetch_live_var = tk.BooleanVar(value=False)
        self.export_csv_var = tk.BooleanVar(value=False)
        self.export_excel_var = tk.BooleanVar(value=False)
        self.csv_path_var = tk.StringVar(value="")
        self.excel_path_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="Ready.")

        self.live_fx_to_usd: dict[str, float] = {}
        self.live_fx_keys: set[str] = set()
        self.live_price_tickers: set[str] = set()

        self._build_ui()
        self.set_rows(len(SAMPLE_PORTFOLIO), SAMPLE_PORTFOLIO)

    def _build_ui(self) -> None:
        root_pad = {"padx": 10, "pady": 8}
        main = ttk.Frame(self.root)
        main.pack(fill="both", expand=True, **root_pad)

        title = ttk.Label(main, text="Portfolio Rebalancer Project", font=("TkDefaultFont", 16, "bold"))
        title.pack(anchor="w")

        subtitle = ttk.Label(
            main,
            text=(
                "Desktop rebalancer with optional live ticker/FX fetch and CSV/Excel export. "
                "Live Yahoo data may be delayed by ~15-20 minutes."
            ),
            foreground="#3f5066",
        )
        subtitle.pack(anchor="w", pady=(0, 8))

        controls = ttk.Frame(main)
        controls.pack(fill="x", pady=(0, 6))

        ttk.Label(controls, text="Number of securities").grid(row=0, column=0, sticky="w")
        self.row_count_spin = tk.Spinbox(
            controls,
            from_=1,
            to=50,
            width=6,
            textvariable=self.row_count_var,
            justify="right",
        )
        self.row_count_spin.grid(row=1, column=0, sticky="w", padx=(0, 10))

        ttk.Label(controls, text="Reporting currency").grid(row=0, column=1, sticky="w")
        self.reporting_currency_combo = ttk.Combobox(
            controls,
            width=10,
            values=list(CURRENCY_OPTIONS.keys()),
            textvariable=self.reporting_currency_var,
            state="readonly",
        )
        self.reporting_currency_combo.grid(row=1, column=1, sticky="w", padx=(0, 10))
        self.reporting_currency_combo.bind("<<ComboboxSelected>>", lambda _e: self.on_reporting_currency_change())

        ttk.Button(controls, text="Apply Rows", command=self.on_apply_rows).grid(row=1, column=2, sticky="w", padx=(0, 6))
        ttk.Button(controls, text="Add Row", command=self.on_add_row).grid(row=1, column=3, sticky="w", padx=(0, 6))
        ttk.Button(controls, text="Remove Last Row", command=self.on_remove_row).grid(row=1, column=4, sticky="w")

        table_section = ttk.LabelFrame(main, text="Portfolio Inputs")
        table_section.pack(fill="both", expand=False, pady=(6, 6))

        header = ttk.Frame(table_section)
        header.pack(fill="x", padx=6, pady=(6, 2))

        header_labels = [
            ("#", 4),
            ("Ticker", 14),
            ("Shares / Units", 14),
            ("Price (local)", 14),
            ("Row Currency", 12),
            ("FX to reporting", 14),
            ("Current Value", 18),
            ("Target Weight %", 14),
        ]

        for col, (label, width) in enumerate(header_labels):
            ttk.Label(header, text=label, width=width, anchor="w").grid(row=0, column=col, padx=2, sticky="w")

        canvas_holder = ttk.Frame(table_section)
        canvas_holder.pack(fill="both", expand=True, padx=6, pady=(0, 4))

        self.table_canvas = tk.Canvas(canvas_holder, height=300, highlightthickness=0)
        self.table_canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(canvas_holder, orient="vertical", command=self.table_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        self.table_canvas.configure(yscrollcommand=scrollbar.set)

        self.rows_frame = ttk.Frame(self.table_canvas)
        self.rows_window_id = self.table_canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")
        self.rows_frame.bind("<Configure>", self._on_rows_frame_configure)
        self.table_canvas.bind("<Configure>", self._on_canvas_configure)

        totals = ttk.Frame(table_section)
        totals.pack(fill="x", padx=6, pady=(2, 8))
        ttk.Label(totals, textvariable=self.live_current_total_var).pack(side="left")
        ttk.Label(totals, textvariable=self.live_weight_total_var).pack(side="right")

        flow_frame = ttk.Frame(main)
        flow_frame.pack(fill="x", pady=(0, 4))
        ttk.Label(flow_frame, textvariable=self.net_flow_label_var).pack(anchor="w")
        self.net_flow_entry = ttk.Entry(flow_frame, textvariable=self.net_flow_var, width=24)
        self.net_flow_entry.pack(anchor="w", pady=(2, 0))
        self.net_flow_entry.bind("<KeyRelease>", lambda _e: self._handle_live_input_change())
        self.net_flow_entry.bind("<FocusOut>", lambda _e: self._handle_live_input_change())

        button_row = ttk.Frame(main)
        button_row.pack(fill="x", pady=(6, 4))
        self.run_button = ttk.Button(button_row, text="Run Rebalance", command=self.on_run_clicked)
        self.run_button.pack(side="left", padx=(0, 6))
        ttk.Button(button_row, text="Load Sample Portfolio", command=self.load_sample_portfolio).pack(side="left")

        options = ttk.LabelFrame(main, text="Run Options")
        options.pack(fill="x", pady=(4, 6))
        options.columnconfigure(1, weight=1)

        self.fetch_live_check = ttk.Checkbutton(
            options,
            text="Fetch live prices and FX rates (requires internet)",
            variable=self.fetch_live_var,
        )
        self.fetch_live_check.grid(row=0, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 3))

        self.csv_check = ttk.Checkbutton(
            options,
            text="Also export to CSV",
            variable=self.export_csv_var,
            command=self._toggle_export_controls,
        )
        self.csv_check.grid(row=1, column=0, sticky="w", padx=6, pady=2)
        self.csv_entry = ttk.Entry(options, textvariable=self.csv_path_var, state="disabled")
        self.csv_entry.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
        self.csv_button = ttk.Button(options, text="Browse...", command=self._browse_csv, state="disabled")
        self.csv_button.grid(row=1, column=2, sticky="ew", padx=6, pady=2)

        self.excel_check = ttk.Checkbutton(
            options,
            text="Also export to Excel",
            variable=self.export_excel_var,
            command=self._toggle_export_controls,
        )
        self.excel_check.grid(row=2, column=0, sticky="w", padx=6, pady=(2, 6))
        self.excel_entry = ttk.Entry(options, textvariable=self.excel_path_var, state="disabled")
        self.excel_entry.grid(row=2, column=1, sticky="ew", padx=4, pady=(2, 6))
        self.excel_button = ttk.Button(options, text="Browse...", command=self._browse_excel, state="disabled")
        self.excel_button.grid(row=2, column=2, sticky="ew", padx=6, pady=(2, 6))

        self.progress = ttk.Progressbar(main, mode="indeterminate")
        self.progress.pack(fill="x", pady=(0, 6))

        self.status_label = tk.Label(main, textvariable=self.status_var, anchor="w", fg=STATUS_NEUTRAL, bg=self.root.cget("bg"))
        self.status_label.pack(fill="x", pady=(0, 2))

        validation_label = ttk.Label(main, textvariable=self.validation_var, foreground="#B12828")
        validation_label.pack(anchor="w", pady=(0, 4))

        output = ttk.LabelFrame(main, text="Rebalance Output")
        output.pack(fill="both", expand=True)
        self.output_text = tk.Text(output, wrap="none", height=15, font=("Courier New", 10))
        self.output_text.pack(fill="both", expand=True, padx=6, pady=6)
        self.output_text.insert("1.0", "Output will appear here after you run rebalancing.")
        self.output_text.configure(state="disabled")

    def _on_rows_frame_configure(self, _event: tk.Event) -> None:
        self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))

    def _on_canvas_configure(self, event: tk.Event) -> None:
        self.table_canvas.itemconfigure(self.rows_window_id, width=event.width)

    def _set_status(self, message: str, color: str = STATUS_NEUTRAL) -> None:
        self.status_var.set(message)
        self.status_label.config(fg=color)

    def _set_running_state(self, running: bool) -> None:
        self.is_running = running
        run_state = "disabled" if running else "normal"
        browse_state = "disabled" if running else "normal"
        self.run_button.config(state=run_state)
        self.row_count_spin.config(state=run_state)
        self.reporting_currency_combo.config(state=run_state)
        self.fetch_live_check.config(state=run_state)
        self.csv_check.config(state=run_state)
        self.excel_check.config(state=run_state)
        self.csv_button.config(state=browse_state if self.export_csv_var.get() else "disabled")
        self.excel_button.config(state=browse_state if self.export_excel_var.get() else "disabled")
        self.csv_entry.config(state=("disabled" if running or not self.export_csv_var.get() else "normal"))
        self.excel_entry.config(state=("disabled" if running or not self.export_excel_var.get() else "normal"))
        if running:
            self.progress.start(12)
        else:
            self.progress.stop()
            self._toggle_export_controls()

    def _render_output(self, text: str) -> None:
        self.output_text.configure(state="normal")
        self.output_text.delete("1.0", "end")
        self.output_text.insert("1.0", text)
        self.output_text.configure(state="disabled")

    def get_reporting_meta(self) -> dict:
        key = self.reporting_currency_var.get()
        return CURRENCY_OPTIONS.get(key, CURRENCY_OPTIONS["USD"])

    def get_row_meta(self, currency_key: str) -> dict:
        return CURRENCY_OPTIONS.get(currency_key, CURRENCY_OPTIONS["USD"])

    def _fallback_fx_to_usd(self, currency_key: str) -> float:
        return float(self.get_row_meta(currency_key)["fx_to_usd"])

    def _get_fx_to_usd(self, currency_key: str) -> float:
        if currency_key in self.live_fx_to_usd:
            return self.live_fx_to_usd[currency_key]
        return self._fallback_fx_to_usd(currency_key)

    def _is_live_fx_pair(self, row_currency: str, report_currency: str) -> bool:
        return row_currency in self.live_fx_keys or report_currency in self.live_fx_keys

    def get_fx_to_reporting(self, currency_key: str) -> float:
        row_to_usd = self._get_fx_to_usd(currency_key)
        report_to_usd = self._get_fx_to_usd(self.reporting_currency_var.get())
        return row_to_usd / report_to_usd

    @staticmethod
    def parse_float(raw: str) -> float | None:
        raw = str(raw).strip()
        if raw == "":
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    @staticmethod
    def format_fx(value: float, is_live: bool = False) -> str:
        return f"~{value:.4f}" if is_live else f"{value:.4f}"

    @staticmethod
    def format_pct(value: float) -> str:
        return f"{value:.2f}%"

    @staticmethod
    def format_shares(value: float) -> str:
        text = f"{value:,.4f}"
        return text.rstrip("0").rstrip(".")

    @staticmethod
    def format_input_price(value: float) -> str:
        text = f"{value:.4f}"
        return text.rstrip("0").rstrip(".")

    def format_currency(self, value: float, currency_key: str | None = None) -> str:
        key = currency_key or self.reporting_currency_var.get()
        meta = self.get_row_meta(key)
        symbol = meta["symbol"]
        if value < 0:
            return f"-{symbol}{abs(value):,.2f}"
        return f"{symbol}{value:,.2f}"

    def _toggle_export_controls(self) -> None:
        if self.is_running:
            self.csv_entry.config(state="disabled")
            self.excel_entry.config(state="disabled")
            self.csv_button.config(state="disabled")
            self.excel_button.config(state="disabled")
            return

        csv_enabled = bool(self.export_csv_var.get())
        excel_enabled = bool(self.export_excel_var.get())
        self.csv_entry.config(state=("normal" if csv_enabled else "disabled"))
        self.csv_button.config(state=("normal" if csv_enabled else "disabled"))
        self.excel_entry.config(state=("normal" if excel_enabled else "disabled"))
        self.excel_button.config(state=("normal" if excel_enabled else "disabled"))

        if not csv_enabled:
            self.csv_path_var.set("")
        if not excel_enabled:
            self.excel_path_var.set("")

    def _browse_csv(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Select CSV output",
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv")],
        )
        if path:
            self.csv_path_var.set(path)

    def _browse_excel(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Select Excel output",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if path:
            self.excel_path_var.set(path)

    def get_current_row_seed(self) -> list[dict]:
        seed = []
        for row in self.rows:
            seed.append(
                {
                    "ticker": row["ticker_var"].get().strip().upper(),
                    "shares": row["shares_var"].get().strip(),
                    "price": row["price_var"].get().strip(),
                    "currencyKey": row["currency_var"].get().strip() or "USD",
                    "targetWeight": row["target_var"].get().strip(),
                }
            )
        return seed

    def clear_live_snapshot(self) -> None:
        self.live_fx_to_usd = {}
        self.live_fx_keys = set()
        self.live_price_tickers = set()

    def set_rows(self, count: int, seed_data: list[dict] | None = None) -> None:
        row_count = clamp_row_count(count)
        self.row_count_var.set(str(row_count))

        for row in self.rows:
            row["frame"].destroy()
        self.rows = []

        for idx in range(row_count):
            data = seed_data[idx] if seed_data and idx < len(seed_data) else {}
            ticker = str(data.get("ticker", "")).upper()
            shares = str(data.get("shares", ""))
            price = str(data.get("price", ""))
            target = str(data.get("targetWeight", ""))
            currency_key = data.get("currencyKey", "USD")
            if currency_key not in CURRENCY_OPTIONS:
                currency_key = "USD"

            frame = ttk.Frame(self.rows_frame)
            frame.grid(row=idx, column=0, sticky="ew", pady=1)
            frame.columnconfigure(1, weight=1)

            ticker_var = tk.StringVar(value=ticker)
            shares_var = tk.StringVar(value=shares)
            price_var = tk.StringVar(value=price)
            target_var = tk.StringVar(value=target)
            currency_var = tk.StringVar(value=currency_key)

            ttk.Label(frame, text=str(idx + 1), width=4).grid(row=0, column=0, padx=2, sticky="w")
            ticker_entry = ttk.Entry(frame, textvariable=ticker_var, width=14)
            ticker_entry.grid(row=0, column=1, padx=2, sticky="w")
            shares_entry = ttk.Entry(frame, textvariable=shares_var, width=14)
            shares_entry.grid(row=0, column=2, padx=2, sticky="w")
            price_entry = ttk.Entry(frame, textvariable=price_var, width=14)
            price_entry.grid(row=0, column=3, padx=2, sticky="w")
            currency_combo = ttk.Combobox(
                frame,
                textvariable=currency_var,
                values=list(CURRENCY_OPTIONS.keys()),
                state="readonly",
                width=12,
            )
            currency_combo.grid(row=0, column=4, padx=2, sticky="w")
            fx_label = ttk.Label(frame, text=self.format_fx(self.get_fx_to_reporting(currency_key)), width=14)
            fx_label.grid(row=0, column=5, padx=2, sticky="w")
            current_label = ttk.Label(frame, text=self.format_currency(0), width=18)
            current_label.grid(row=0, column=6, padx=2, sticky="w")
            target_entry = ttk.Entry(frame, textvariable=target_var, width=14)
            target_entry.grid(row=0, column=7, padx=2, sticky="w")

            for entry in (ticker_entry, shares_entry, price_entry, target_entry):
                entry.bind("<KeyRelease>", lambda _e: self._handle_live_input_change())
                entry.bind("<FocusOut>", lambda _e: self._handle_live_input_change())

            currency_combo.bind("<<ComboboxSelected>>", lambda _e: self._handle_live_input_change())

            self.rows.append(
                {
                    "frame": frame,
                    "ticker_var": ticker_var,
                    "shares_var": shares_var,
                    "price_var": price_var,
                    "target_var": target_var,
                    "currency_var": currency_var,
                    "fx_label": fx_label,
                    "current_label": current_label,
                }
            )

        self.update_currency_ui()
        self.update_live_totals()

    def update_currency_ui(self) -> None:
        label = self.get_reporting_meta()["label"]
        self.net_flow_label_var.set(f"Net contribution / withdrawal ({label})")

    def _handle_live_input_change(self) -> None:
        self.validation_var.set("")
        self.update_live_totals()
        if self.has_calculated and not self.is_running:
            self.recalculate_from_inputs(show_error=False)

    def update_live_totals(self) -> None:
        total_current = 0.0
        total_weight = 0.0
        report_key = self.reporting_currency_var.get()

        for row in self.rows:
            shares = self.parse_float(row["shares_var"].get())
            price = self.parse_float(row["price_var"].get())
            target = self.parse_float(row["target_var"].get())
            currency_key = row["currency_var"].get()
            fx = self.get_fx_to_reporting(currency_key)
            fx_is_live = self._is_live_fx_pair(currency_key, report_key)

            if shares is not None and price is not None and shares >= 0 and price > 0:
                local_value = shares * price
                current_value = local_value * fx
            else:
                current_value = 0.0

            row["fx_label"].configure(text=self.format_fx(fx, is_live=fx_is_live))
            row["current_label"].configure(text=self.format_currency(current_value))

            total_current += current_value
            if target is not None and target >= 0:
                total_weight += target

        self.live_current_total_var.set(f"Current total: {self.format_currency(total_current)}")
        self.live_weight_total_var.set(f"Target weight total: {self.format_pct(total_weight)}")

    def parse_positions(self, require_price: bool = True) -> list[dict]:
        if not self.rows:
            raise ValueError("Add at least one security row.")

        positions = []
        for idx, row in enumerate(self.rows):
            ticker = row["ticker_var"].get().strip().upper()
            shares = self.parse_float(row["shares_var"].get())
            price = self.parse_float(row["price_var"].get())
            target_weight = self.parse_float(row["target_var"].get())
            currency_key = row["currency_var"].get()

            if not ticker:
                raise ValueError(f"Row {idx + 1}: ticker is required.")
            if shares is None or shares < 0:
                raise ValueError(f"Row {idx + 1}: shares/units must be a non-negative number.")
            if require_price and (price is None or price <= 0):
                raise ValueError(f"Row {idx + 1}: price must be greater than 0.")
            if not require_price and price is not None and price <= 0:
                raise ValueError(f"Row {idx + 1}: manual price must be greater than 0 when provided.")
            if target_weight is None or target_weight < 0:
                raise ValueError(f"Row {idx + 1}: target weight must be 0 or greater.")
            if currency_key not in CURRENCY_OPTIONS:
                raise ValueError(f"Row {idx + 1}: row currency is invalid.")

            positions.append(
                {
                    "rowIndex": idx,
                    "ticker": ticker,
                    "shares": shares,
                    "price": price,
                    "currencyKey": currency_key,
                    "currencyLabel": self.get_row_meta(currency_key)["label"],
                    "targetWeight": target_weight,
                }
            )

        if not any(position["targetWeight"] > 0 for position in positions):
            raise ValueError("At least one target weight must be greater than 0.")

        return positions

    def _validate_export_paths(self) -> tuple[Path | None, Path | None]:
        csv_path: Path | None = None
        excel_path: Path | None = None

        if self.export_csv_var.get():
            csv_raw = self.csv_path_var.get().strip()
            if not csv_raw:
                raise ValueError("CSV export is enabled but no CSV file path is selected.")
            csv_path = Path(csv_raw)
            if csv_path.suffix.lower() != ".csv":
                raise ValueError("CSV export path must end with .csv")
            if not csv_path.parent.exists():
                raise ValueError("CSV export folder does not exist.")

        if self.export_excel_var.get():
            excel_raw = self.excel_path_var.get().strip()
            if not excel_raw:
                raise ValueError("Excel export is enabled but no .xlsx path is selected.")
            excel_path = Path(excel_raw)
            if excel_path.suffix.lower() != ".xlsx":
                raise ValueError("Excel export path must end with .xlsx")
            if not excel_path.parent.exists():
                raise ValueError("Excel export folder does not exist.")
            if Workbook is None:
                raise ValueError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        return csv_path, excel_path

    def _attach_fx_values(self, positions: list[dict], fx_to_usd_map: dict[str, float]) -> list[dict]:
        reporting_key = self.reporting_currency_var.get()
        report_to_usd = fx_to_usd_map.get(reporting_key, self._fallback_fx_to_usd(reporting_key))

        enriched = []
        for position in positions:
            currency_key = position["currencyKey"]
            row_to_usd = fx_to_usd_map.get(currency_key, self._fallback_fx_to_usd(currency_key))
            fx_to_reporting = row_to_usd / report_to_usd
            price = position["price"]
            if price is None or price <= 0:
                raise ValueError(f"{position['ticker']}: price must be greater than 0.")
            current_value_local = position["shares"] * price
            enriched.append(
                {
                    **position,
                    "price": price,
                    "fxToReporting": fx_to_reporting,
                    "currentValueLocal": current_value_local,
                    "currentValue": current_value_local * fx_to_reporting,
                }
            )
        return enriched

    def build_output(
        self,
        summary: dict,
        rows: list[dict],
        warnings: list[str] | None = None,
        exports: list[str] | None = None,
        live_data_used: bool = False,
    ) -> str:
        lines = []
        lines.append("SUMMARY")
        lines.append("-" * 96)
        lines.append(f"Run Date:            {summary['runDate']}")
        lines.append(f"Live data used:      {'Yes' if live_data_used else 'No'}")
        lines.append(f"Total Current Value: {self.format_currency(summary['totalCurrent'])}")
        lines.append(f"Net Flow:            {self.format_currency(summary['netFlow'])}")
        lines.append(f"Target Ending Value: {self.format_currency(summary['endingValue'])}")
        lines.append(f"Total Buy Value:     {self.format_currency(summary['totalBuys'])}")
        lines.append(f"Total Sell Value:    {self.format_currency(summary['totalSells'])}")
        lines.append("")
        lines.append("TRADE PLAN")
        lines.append("-" * 170)
        lines.append(
            f"{'Ticker':<8}{'Curr':<10}{'Shares':>12}{'Price':>14}{'Current':>14}"
            f"{'Target%':>10}{'Target':>14}{'Trade':>14}{'Trade Local':>14}"
            f"{'Trade Sh':>12}{'Action':>9}{'Post Sh':>12}"
        )
        lines.append("-" * 170)
        for item in rows:
            lines.append(
                f"{item['ticker']:<8}{item['currencyLabel']:<10}"
                f"{self.format_shares(item['shares']):>12}"
                f"{self.format_currency(item['price'], item['currencyKey']):>14}"
                f"{self.format_currency(item['currentValue']):>14}"
                f"{self.format_pct(item['targetWeightNorm'] * 100):>10}"
                f"{self.format_currency(item['targetValue']):>14}"
                f"{self.format_currency(item['tradeValue']):>14}"
                f"{self.format_currency(item['tradeValueLocal'], item['currencyKey']):>14}"
                f"{self.format_shares(item['tradeShares']):>12}"
                f"{item['action']:>9}"
                f"{self.format_shares(item['postTradeShares']):>12}"
            )

        if exports:
            lines.append("")
            lines.append("EXPORTS")
            lines.append("-" * 96)
            for line in exports:
                lines.append(f"- {line}")

        if warnings:
            lines.append("")
            lines.append("WARNINGS")
            lines.append("-" * 96)
            for warning in warnings:
                lines.append(f"- {warning}")

        return "\n".join(lines)

    def recalculate_from_inputs(self, show_error: bool = True) -> None:
        self.validation_var.set("")
        try:
            positions = self.parse_positions(require_price=True)
            net_flow = self.parse_float(self.net_flow_var.get())
            if net_flow is None:
                raise ValueError("Net contribution / withdrawal must be a valid number.")

            fx_map = {key: self._get_fx_to_usd(key) for key in CURRENCY_OPTIONS}
            enriched_positions = self._attach_fx_values(positions, fx_map)
            summary, results = calculate_rebalance_plan(enriched_positions, net_flow)
            self._render_output(self.build_output(summary, results, live_data_used=bool(self.live_fx_keys)))
            self.has_calculated = True
        except ValueError as exc:
            self._render_output("Output will appear here after you run rebalancing.")
            self.validation_var.set(str(exc))
            self.has_calculated = False
            if show_error:
                self.root.bell()

    def on_run_clicked(self) -> None:
        if self.is_running:
            return

        self.validation_var.set("")
        wants_live_fetch = bool(self.fetch_live_var.get())
        use_live_fetch = wants_live_fetch and yf is not None
        pre_warnings: list[str] = []
        if wants_live_fetch and not use_live_fetch:
            pre_warnings.append(
                "yfinance is not installed; using manual prices and fallback FX rates. "
                "Install with: pip install yfinance"
            )

        try:
            csv_path, excel_path = self._validate_export_paths()
            positions = self.parse_positions(require_price=not use_live_fetch)
            net_flow = self.parse_float(self.net_flow_var.get())
            if net_flow is None:
                raise ValueError("Net contribution / withdrawal must be a valid number.")
        except ValueError as exc:
            self.validation_var.set(str(exc))
            self.root.bell()
            return

        self._set_running_state(True)
        if use_live_fetch:
            self._set_status("Fetching live prices and FX rates...", STATUS_NEUTRAL)
        else:
            self._set_status("Running rebalance...", STATUS_NEUTRAL)

        run_payload = {
            "positions": positions,
            "net_flow": net_flow,
            "use_live_fetch": use_live_fetch,
            "pre_warnings": pre_warnings,
            "csv_path": csv_path,
            "excel_path": excel_path,
            "reporting_currency": self.reporting_currency_var.get(),
        }

        self.worker_thread = threading.Thread(target=self._run_worker, args=(run_payload,), daemon=True)
        self.worker_thread.start()
        self.root.after(80, self._poll_worker_events)

    def _run_worker(self, payload: dict) -> None:
        try:
            warnings = list(payload["pre_warnings"])
            positions = [dict(position) for position in payload["positions"]]
            net_flow = float(payload["net_flow"])
            reporting_currency = str(payload["reporting_currency"])
            use_live_fetch = bool(payload["use_live_fetch"])
            csv_path: Path | None = payload["csv_path"]
            excel_path: Path | None = payload["excel_path"]

            fx_to_usd_map = {key: self._fallback_fx_to_usd(key) for key in CURRENCY_OPTIONS}
            live_fx_keys: set[str] = set()
            live_price_tickers: set[str] = set()
            live_price_updates: dict[int, float] = {}
            live_currency_updates: dict[int, str] = {}

            if use_live_fetch:
                self.worker_events.put(("status", "Fetching live prices from Yahoo Finance..."))
                for position in positions:
                    ticker = position["ticker"]
                    row_idx = int(position["rowIndex"])
                    selected_currency_key = position["currencyKey"]
                    selected_currency_code = CURRENCY_OPTIONS[selected_currency_key]["code"]
                    manual_price = position["price"]

                    best_matching_quote = None
                    first_fallback_quote = None
                    for candidate in _build_ticker_candidates(ticker, selected_currency_key):
                        quote_details = fetch_live_quote_details(candidate)
                        if quote_details is None:
                            continue
                        quote_code = quote_details.get("quoteCurrency")
                        if quote_code == selected_currency_code:
                            best_matching_quote = quote_details
                            break
                        if first_fallback_quote is None:
                            first_fallback_quote = quote_details

                    chosen_quote = best_matching_quote or first_fallback_quote
                    if chosen_quote is None:
                        if manual_price is None:
                            raise ValueError(
                                f"Live price unavailable for {ticker} and no manual price is provided."
                            )
                        warnings.append(
                            f"Live price unavailable for {ticker}; using manual entry {manual_price:.4f}."
                        )
                        continue

                    resolved_ticker = str(chosen_quote["resolvedTicker"])
                    quote_code = chosen_quote.get("quoteCurrency")
                    quote_key = CURRENCY_CODE_TO_KEY.get(quote_code) if quote_code else None
                    chosen_price = float(chosen_quote["price"])

                    if quote_code and quote_code != selected_currency_code:
                        if quote_key is not None and manual_price is None:
                            position["currencyKey"] = quote_key
                            position["currencyLabel"] = self.get_row_meta(quote_key)["label"]
                            live_currency_updates[row_idx] = quote_key
                            position["price"] = chosen_price
                            live_price_updates[row_idx] = chosen_price
                            live_price_tickers.add(ticker)
                            warnings.append(
                                f"{ticker}: live quote currency {quote_code} does not match selected "
                                f"{selected_currency_code}; adjusted row currency to {quote_code} for this run."
                            )
                            if resolved_ticker != ticker:
                                warnings.append(
                                    f"{ticker}: used exchange symbol {resolved_ticker} for live quote."
                                )
                            continue

                        if manual_price is not None:
                            warnings.append(
                                f"{ticker}: live quote currency {quote_code} does not match selected "
                                f"{selected_currency_code}; keeping manual entry {manual_price:.4f}."
                            )
                            continue

                        raise ValueError(
                            f"{ticker}: live quote returned unsupported quote currency {quote_code} "
                            f"for selected {selected_currency_code}; enter a manual price."
                        )

                    position["price"] = chosen_price
                    live_price_updates[row_idx] = chosen_price
                    live_price_tickers.add(ticker)
                    if resolved_ticker != ticker:
                        warnings.append(f"{ticker}: used exchange symbol {resolved_ticker} for live quote.")
                    if chosen_quote.get("unitScale", 1.0) != 1.0:
                        warnings.append(
                            f"{ticker}: converted sub-unit quote currency "
                            f"{chosen_quote.get('rawQuoteCurrency')} to {quote_code} units."
                        )

                self.worker_events.put(("status", "Fetching live FX rates..."))
                currencies_in_scope = {reporting_currency}
                currencies_in_scope.update(position["currencyKey"] for position in positions)
                for currency_key in sorted(currencies_in_scope):
                    code = CURRENCY_OPTIONS[currency_key]["code"]
                    live_rate = fetch_live_fx_to_usd(code)
                    if live_rate is not None:
                        fx_to_usd_map[currency_key] = live_rate
                        live_fx_keys.add(currency_key)
                    else:
                        fallback = self._fallback_fx_to_usd(currency_key)
                        warnings.append(
                            f"Live FX unavailable for {currency_key}; using fallback rate {fallback:.4f}."
                        )

            enriched_positions = self._attach_fx_values(positions, fx_to_usd_map)
            summary, results = calculate_rebalance_plan(enriched_positions, net_flow)

            export_lines: list[str] = []
            if csv_path is not None:
                self.worker_events.put(("status", "Writing CSV export..."))
                write_csv_export(csv_path, summary, results, reporting_currency)
                export_lines.append(f"CSV: {csv_path}")

            if excel_path is not None:
                self.worker_events.put(("status", "Writing Excel export..."))
                write_excel_export(excel_path, summary, results, reporting_currency)
                export_lines.append(f"Excel: {excel_path}")

            self.worker_events.put(
                (
                    "success",
                    {
                        "summary": summary,
                        "results": results,
                        "warnings": warnings,
                        "exports": export_lines,
                        "live_data_used": use_live_fetch,
                        "live_fx_to_usd": fx_to_usd_map if use_live_fetch else {},
                        "live_fx_keys": live_fx_keys,
                        "live_price_tickers": live_price_tickers,
                        "live_price_updates": live_price_updates,
                        "live_currency_updates": live_currency_updates,
                    },
                )
            )
        except Exception as exc:  # pylint: disable=broad-except
            self.worker_events.put(("error", str(exc)))

    def _poll_worker_events(self) -> None:
        while True:
            try:
                event, payload = self.worker_events.get_nowait()
            except queue.Empty:
                break

            if event == "status":
                self._set_status(str(payload), STATUS_NEUTRAL)
            elif event == "success":
                data = payload
                assert isinstance(data, dict)
                self._set_running_state(False)
                self.live_fx_to_usd = dict(data["live_fx_to_usd"])
                self.live_fx_keys = set(data["live_fx_keys"])
                self.live_price_tickers = set(data["live_price_tickers"])
                for row_idx, new_price in data.get("live_price_updates", {}).items():
                    idx = int(row_idx)
                    if 0 <= idx < len(self.rows):
                        self.rows[idx]["price_var"].set(self.format_input_price(float(new_price)))
                for row_idx, new_currency_key in data.get("live_currency_updates", {}).items():
                    idx = int(row_idx)
                    if 0 <= idx < len(self.rows):
                        self.rows[idx]["currency_var"].set(new_currency_key)
                self.update_live_totals()
                output_text = self.build_output(
                    summary=data["summary"],
                    rows=data["results"],
                    warnings=data["warnings"],
                    exports=data["exports"],
                    live_data_used=bool(data["live_data_used"]),
                )
                self._render_output(output_text)
                self.has_calculated = True
                if data["warnings"]:
                    self._set_status(
                        f"Completed with {len(data['warnings'])} warning(s). Verify data before trading.",
                        STATUS_WARN,
                    )
                else:
                    self._set_status("Rebalance completed successfully.", STATUS_SUCCESS)
            elif event == "error":
                self._set_running_state(False)
                self._render_output("Output will appear here after you run rebalancing.")
                self.validation_var.set(str(payload))
                self.has_calculated = False
                self._set_status("Run failed. Review validation message.", STATUS_ERROR)
                self.root.bell()

        if self.is_running:
            self.root.after(80, self._poll_worker_events)

    def on_apply_rows(self) -> None:
        seed = self.get_current_row_seed()
        try:
            requested = int(float(self.row_count_var.get() or "1"))
        except ValueError:
            requested = 1
        count = clamp_row_count(requested)
        trimmed = [seed[i] if i < len(seed) else {} for i in range(count)]
        self.set_rows(count, trimmed)

    def on_add_row(self) -> None:
        seed = self.get_current_row_seed()
        if len(seed) >= 50:
            self.validation_var.set("Maximum row limit reached (50).")
            return
        seed.append({})
        self.set_rows(len(seed), seed)

    def on_remove_row(self) -> None:
        seed = self.get_current_row_seed()
        if len(seed) <= 1:
            self.validation_var.set("At least one row is required.")
            return
        seed.pop()
        self.set_rows(len(seed), seed)

    def on_reporting_currency_change(self) -> None:
        self.validation_var.set("")
        self.update_currency_ui()
        self.update_live_totals()
        if self.has_calculated and not self.is_running:
            self.recalculate_from_inputs(show_error=False)

    def load_sample_portfolio(self) -> None:
        self.clear_live_snapshot()
        self.row_count_var.set(str(len(SAMPLE_PORTFOLIO)))
        self.set_rows(len(SAMPLE_PORTFOLIO), SAMPLE_PORTFOLIO)
        self.net_flow_var.set("0")
        self.validation_var.set("")
        self._set_status("Sample portfolio loaded.", STATUS_NEUTRAL)
        self._render_output("Output will appear here after you run rebalancing.")
        self.has_calculated = False


def main() -> None:
    if not TK_AVAILABLE:
        raise RuntimeError(
            "tkinter is not available in this Python environment. "
            "Install tkinter support to run the desktop GUI."
        )
    root = tk.Tk()
    PortfolioRebalancerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
