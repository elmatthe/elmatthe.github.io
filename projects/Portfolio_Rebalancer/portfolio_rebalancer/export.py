"""CSV and Excel export for rebalancer results."""
from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path

from portfolio_rebalancer.fx import (
    EXCEL_CURRENCY_FORMATS,
    format_currency,
    format_pct,
    format_shares,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]

BRAND_NAVY_ARGB = "FF1A2E4A"


def _replace_or_create_sheet(workbook: "Workbook", title: str):  # type: ignore[type-arg]
    if title in workbook.sheetnames:
        idx = workbook.sheetnames.index(title)
        del workbook[title]
        return workbook.create_sheet(title=title, index=idx)
    return workbook.create_sheet(title=title)


def write_csv_export(
    csv_path: Path,
    summary: dict,
    rows: list[dict],
    reporting_currency: str,
    warnings: list[str] | None = None,
    notes: list[str] | None = None,
) -> None:
    run_date = dt.datetime.now().isoformat(timespec="seconds")
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Run Date", run_date])
        w.writerow(["Mode", summary.get("mode", "")])
        w.writerow(["Reporting Currency", reporting_currency])
        w.writerow(["Total Current Value", f"{summary['totalCurrent']:.2f}"])
        if summary.get("mode") == "new_money":
            w.writerow(["New Money Budget", f"{summary.get('budget', 0):.2f}"])
        w.writerow(["Target Pool", f"{summary['pool']:.2f}"])
        w.writerow(["Total Buy Value", f"{summary['totalBuys']:.2f}"])
        w.writerow(["Total Sell Value", f"{summary['totalSells']:.2f}"])
        w.writerow([])
        headers = [
            "Ticker", "Account Type", "Currency", "Shares", "Price (Local)",
            "Current Value", "Target %", "Target Value", "Trade (Reporting)",
            "Trade (Local)", "Trade Shares", "Action", "Post-Trade Shares",
        ]
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r["ticker"], r.get("accountType", ""), r["currencyLabel"],
                f"{r['shares']:.4f}", f"{r['price']:.4f}",
                f"{r['currentValue']:.2f}", f"{r['targetWeightNorm'] * 100:.2f}",
                f"{r['targetValue']:.2f}", f"{r['tradeValue']:.2f}",
                f"{r['tradeValueLocal']:.2f}", f"{r['tradeShares']:.4f}",
                r["action"], f"{r['postTradeShares']:.4f}",
            ])
        if warnings:
            w.writerow([])
            w.writerow(["Warnings"])
            for warning in warnings:
                w.writerow([warning])
        if notes:
            w.writerow([])
            w.writerow(["Notes"])
            for note in notes:
                w.writerow([note])


def write_excel_export(
    excel_path: Path,
    summary: dict,
    rows: list[dict],
    reporting_currency: str,
) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")
    assert Font is not None and PatternFill is not None and Alignment is not None

    is_new = not excel_path.exists()
    wb = load_workbook(excel_path) if excel_path.exists() else Workbook()
    summary_ws = _replace_or_create_sheet(wb, "RB_Summary")
    trades_ws = _replace_or_create_sheet(wb, "RB_TradePlan")
    if is_new and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    navy = PatternFill(fill_type="solid", start_color=BRAND_NAVY_ARGB, end_color=BRAND_NAVY_ARGB)
    white_bold = Font(color="FFFFFF", bold=True)
    heading = Font(bold=True)
    rep_fmt = EXCEL_CURRENCY_FORMATS.get(reporting_currency, '"$"#,##0.00')

    summary_ws.merge_cells("A1:B1")
    h = summary_ws["A1"]
    h.value = "Portfolio Rebalancer Summary"
    h.fill = navy
    h.font = white_bold
    h.alignment = Alignment(horizontal="center")

    run_date = dt.datetime.now().isoformat(timespec="seconds")
    mode_label = "New Money" if summary.get("mode") == "new_money" else "Rebalance (Pure)"
    info_rows = [
        (3, "Run Date", run_date, None),
        (4, "Mode", mode_label, None),
        (5, "Reporting Currency", reporting_currency, None),
        (6, "Total Current Value", summary["totalCurrent"], rep_fmt),
    ]
    if summary.get("mode") == "new_money":
        info_rows.append((7, "New Money Budget", summary.get("budget", 0), rep_fmt))
        info_rows.append((8, "Target Pool", summary["pool"], rep_fmt))
        info_rows.append((9, "Total Buy Value", summary["totalBuys"], rep_fmt))
        info_rows.append((10, "Total Sell Value", summary["totalSells"], rep_fmt))
    else:
        info_rows.append((7, "Target Pool", summary["pool"], rep_fmt))
        info_rows.append((8, "Total Buy Value", summary["totalBuys"], rep_fmt))
        info_rows.append((9, "Total Sell Value", summary["totalSells"], rep_fmt))

    for row_idx, label, value, nf in info_rows:
        summary_ws.cell(row=row_idx, column=1, value=label).font = heading
        c = summary_ws.cell(row=row_idx, column=2, value=value)
        if nf:
            c.number_format = nf
    summary_ws.column_dimensions["A"].width = 34
    summary_ws.column_dimensions["B"].width = 26

    trade_headers = [
        "Ticker", "Account Type", "Currency", "Shares", "Price (Local)",
        "Current Value", "Target %", "Target Value", "Trade (Reporting)",
        "Trade (Local)", "Trade Shares", "Action", "Post-Trade Shares",
    ]
    for col_idx, label in enumerate(trade_headers, start=1):
        c = trades_ws.cell(row=1, column=col_idx, value=label)
        c.fill = navy
        c.font = white_bold

    for ri, r in enumerate(rows, start=2):
        trades_ws.cell(row=ri, column=1, value=r["ticker"])
        trades_ws.cell(row=ri, column=2, value=r.get("accountType", ""))
        trades_ws.cell(row=ri, column=3, value=r["currencyLabel"])
        trades_ws.cell(row=ri, column=4, value=float(r["shares"])).number_format = "#,##0.0000"
        trades_ws.cell(row=ri, column=5, value=float(r["price"])).number_format = "#,##0.0000"
        trades_ws.cell(row=ri, column=6, value=float(r["currentValue"])).number_format = rep_fmt
        trades_ws.cell(row=ri, column=7, value=float(r["targetWeightNorm"])).number_format = "0.00%"
        trades_ws.cell(row=ri, column=8, value=float(r["targetValue"])).number_format = rep_fmt
        trades_ws.cell(row=ri, column=9, value=float(r["tradeValue"])).number_format = rep_fmt
        trades_ws.cell(row=ri, column=10, value=float(r["tradeValueLocal"])).number_format = "#,##0.00"
        trades_ws.cell(row=ri, column=11, value=float(r["tradeShares"])).number_format = "#,##0.0000"
        ac = trades_ws.cell(row=ri, column=12, value=r["action"])
        trades_ws.cell(row=ri, column=13, value=float(r["postTradeShares"])).number_format = "#,##0.0000"
        if r["action"] == "Buy":
            ac.fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        elif r["action"] == "Sell":
            ac.fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    widths = [12, 16, 12, 14, 14, 14, 12, 14, 16, 14, 14, 10, 16]
    for i, w in enumerate(widths):
        trades_ws.column_dimensions[chr(65 + i)].width = w
    trades_ws.freeze_panes = "A2"

    try:
        wb.save(excel_path)
    except PermissionError as exc:
        raise PermissionError(
            "The workbook appears to be open. Close it in Excel and try again."
        ) from exc
