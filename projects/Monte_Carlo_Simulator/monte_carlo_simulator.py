"""Monte Carlo Retirement Simulator desktop application.

This version includes:
- Full input validation
- Threaded simulation execution with a live status bar
- Excel report writing (summary, percentiles, chart)
- Optional CSV export of percentile data
"""

from __future__ import annotations

import csv
import datetime as dt
import queue
import tempfile
import threading
from dataclasses import dataclass
from pathlib import Path

TK_AVAILABLE = True
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    TK_AVAILABLE = False
    tk = None  # type: ignore[assignment]
    filedialog = None  # type: ignore[assignment]
    messagebox = None  # type: ignore[assignment]
    ttk = None  # type: ignore[assignment]

MISSING_DEPENDENCIES: list[str] = []

try:
    import numpy as np
except ImportError:
    np = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("numpy")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    XLImage = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("openpyxl")

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter
except ImportError:
    plt = None  # type: ignore[assignment]
    FuncFormatter = None  # type: ignore[assignment]
    MISSING_DEPENDENCIES.append("matplotlib")


WINDOW_TITLE = "Monte Carlo Retirement Simulator"
WINDOW_SIZE = "480x580"
BRAND_NAVY = "#1a2e4a"
BRAND_NAVY_ARGB = "FF1A2E4A"
STATUS_NEUTRAL = "#1f4f8a"
STATUS_ERROR = "#9b2f2f"
STATUS_SUCCESS = "#1c7550"


@dataclass
class SimulationInputs:
    current_portfolio: float
    annual_contribution: float
    contribution_growth_rate: float
    years_to_retirement: int
    years_in_retirement: int
    expected_return: float
    volatility: float
    inflation_rate: float
    annual_spending: float
    pension_income: float
    simulations: int
    workbook_path: Path
    csv_path: Path | None = None


@dataclass
class SimulationResult:
    years: "np.ndarray"
    p10: "np.ndarray"
    p25: "np.ndarray"
    p50: "np.ndarray"
    p75: "np.ndarray"
    p90: "np.ndarray"
    success_probability: float
    median_portfolio_at_retirement: float
    median_final_value: float
    p10_final_value: float
    p90_final_value: float
    safe_withdrawal_rate: float
    failed_simulations: int
    median_ruin_year: float | None
    net_withdrawal: float


class ValidationError(Exception):
    """Raised when user input validation fails."""


def _require_dependencies() -> None:
    if MISSING_DEPENDENCIES:
        deps = ", ".join(MISSING_DEPENDENCIES)
        raise RuntimeError(
            f"Missing required dependencies: {deps}. "
            "Install them with: pip install openpyxl matplotlib numpy"
        )


def _add_years_safe(base_date: dt.date, years: int) -> dt.date:
    target_year = base_date.year + years
    try:
        return base_date.replace(year=target_year)
    except ValueError:
        return base_date.replace(year=target_year, day=28)


def _replace_or_create_sheet(workbook: "Workbook", title: str):
    if title in workbook.sheetnames:
        idx = workbook.sheetnames.index(title)
        del workbook[title]
        sheet = workbook.create_sheet(title=title, index=idx)
    else:
        sheet = workbook.create_sheet(title=title)
    return sheet


def run_monte_carlo(inputs: SimulationInputs) -> SimulationResult:
    _require_dependencies()
    assert np is not None

    total_years = inputs.years_to_retirement + inputs.years_in_retirement
    years = np.arange(total_years + 1)

    paths = np.zeros((inputs.simulations, total_years + 1), dtype=float)
    ruin_years: list[int] = []
    final_values = np.zeros(inputs.simulations, dtype=float)
    retirement_values = np.zeros(inputs.simulations, dtype=float)

    net_withdrawal = inputs.annual_spending - inputs.pension_income
    mean_return = inputs.expected_return / 100.0
    volatility = inputs.volatility / 100.0
    contrib_growth = inputs.contribution_growth_rate / 100.0

    for sim_idx in range(inputs.simulations):
        portfolio = inputs.current_portfolio
        contribution = inputs.annual_contribution
        paths[sim_idx, 0] = portfolio
        ruined = False
        ruin_year = None

        for year in range(1, inputs.years_to_retirement + 1):
            annual_return = float(np.random.normal(mean_return, volatility))
            portfolio = portfolio * (1.0 + annual_return) + contribution
            if portfolio < 0:
                portfolio = 0.0
            paths[sim_idx, year] = portfolio
            contribution *= 1.0 + contrib_growth

        for retire_year in range(1, inputs.years_in_retirement + 1):
            year_idx = inputs.years_to_retirement + retire_year

            if ruined:
                paths[sim_idx, year_idx] = 0.0
                continue

            annual_return = float(np.random.normal(mean_return, volatility))
            portfolio = portfolio * (1.0 + annual_return) - net_withdrawal
            if portfolio <= 0:
                portfolio = 0.0
                ruined = True
                ruin_year = retire_year
            paths[sim_idx, year_idx] = portfolio

        if ruin_year is not None:
            ruin_years.append(ruin_year)

        final_values[sim_idx] = portfolio
        retirement_values[sim_idx] = paths[sim_idx, inputs.years_to_retirement]

    p10, p25, p50, p75, p90 = np.percentile(paths, [10, 25, 50, 75, 90], axis=0)

    failed_simulations = len(ruin_years)
    success_probability = ((inputs.simulations - failed_simulations) / inputs.simulations) * 100.0
    median_retirement = float(np.percentile(retirement_values, 50))
    median_final = float(np.percentile(final_values, 50))
    p10_final = float(np.percentile(final_values, 10))
    p90_final = float(np.percentile(final_values, 90))
    swr = (net_withdrawal / median_retirement * 100.0) if median_retirement > 0 else 0.0
    median_ruin_year = float(np.median(np.array(ruin_years))) if ruin_years else None

    return SimulationResult(
        years=years,
        p10=p10,
        p25=p25,
        p50=p50,
        p75=p75,
        p90=p90,
        success_probability=success_probability,
        median_portfolio_at_retirement=median_retirement,
        median_final_value=median_final,
        p10_final_value=p10_final,
        p90_final_value=p90_final,
        safe_withdrawal_rate=swr,
        failed_simulations=failed_simulations,
        median_ruin_year=median_ruin_year,
        net_withdrawal=net_withdrawal,
    )


def _write_summary_sheet(sheet, inputs: SimulationInputs, result: SimulationResult) -> None:
    assert Font is not None and PatternFill is not None and Alignment is not None

    navy_fill = PatternFill(fill_type="solid", start_color=BRAND_NAVY_ARGB, end_color=BRAND_NAVY_ARGB)
    white_bold = Font(color="FFFFFF", bold=True)
    heading_font = Font(bold=True)

    sheet.merge_cells("A1:B1")
    header = sheet["A1"]
    header.value = "Monte Carlo Simulation Summary"
    header.fill = navy_fill
    header.font = white_bold
    header.alignment = Alignment(horizontal="center")

    rows = [
        (3, "Run Date", dt.date.today().isoformat()),
        (4, "Simulations Run", inputs.simulations),
        (5, "Current Portfolio", inputs.current_portfolio),
        (6, "Annual Contribution", inputs.annual_contribution),
        (7, "Contribution Growth Rate", inputs.contribution_growth_rate / 100.0),
        (8, "Years to Retirement", inputs.years_to_retirement),
        (9, "Years in Retirement", inputs.years_in_retirement),
        (10, "Expected Return", inputs.expected_return / 100.0),
        (11, "Volatility", inputs.volatility / 100.0),
        (12, "Inflation Rate", inputs.inflation_rate / 100.0),
        (13, "Annual Retirement Spending", inputs.annual_spending),
        (14, "CPP / OAS / Pension Income", inputs.pension_income),
        (15, "Net Annual Withdrawal", result.net_withdrawal),
    ]

    for row_idx, label, value in rows:
        sheet.cell(row=row_idx, column=1, value=label).font = heading_font
        cell = sheet.cell(row=row_idx, column=2, value=value)
        if row_idx in {5, 6, 13, 14, 15}:
            cell.number_format = "$#,##0.00"
        elif row_idx in {7, 10, 11, 12}:
            cell.number_format = "0.00%"

    sheet.cell(row=17, column=1, value="RESULTS").font = Font(bold=True, color=BRAND_NAVY_ARGB)

    summary_rows = [
        (18, "Probability of Success", result.success_probability / 100.0, "0.0%"),
        (19, "Median Portfolio at Retirement", result.median_portfolio_at_retirement, "$#,##0.00"),
        (20, "Median Final Portfolio Value", result.median_final_value, "$#,##0.00"),
        (21, "10th Percentile Final Value", result.p10_final_value, "$#,##0.00"),
        (22, "90th Percentile Final Value", result.p90_final_value, "$#,##0.00"),
        (23, "Safe Withdrawal Rate", result.safe_withdrawal_rate / 100.0, "0.00%"),
        (
            24,
            "Failed Simulations",
            f"{result.failed_simulations} of {inputs.simulations}",
            None,
        ),
        (
            25,
            "Median Ruin Year (if any)",
            (
                f"Year {result.median_ruin_year:.1f} of retirement"
                if result.median_ruin_year is not None
                else "N/A - no failures"
            ),
            None,
        ),
    ]

    for row_idx, label, value, number_format in summary_rows:
        sheet.cell(row=row_idx, column=1, value=label).font = heading_font
        val_cell = sheet.cell(row=row_idx, column=2, value=value)
        if number_format:
            val_cell.number_format = number_format

    probability_cell = sheet.cell(row=18, column=2)
    if result.success_probability >= 85:
        probability_cell.fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    elif result.success_probability >= 70:
        probability_cell.fill = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
    else:
        probability_cell.fill = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")

    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 26


def _write_percentiles_sheet(sheet, inputs: SimulationInputs, result: SimulationResult) -> None:
    assert Font is not None and PatternFill is not None

    headers = ["Year", "P10", "P25", "P50 (Median)", "P75", "P90", "Phase marker"]
    header_fill = PatternFill(fill_type="solid", start_color=BRAND_NAVY_ARGB, end_color=BRAND_NAVY_ARGB)
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for idx, year in enumerate(result.years, start=2):
        year_i = int(year)
        sheet.cell(row=idx, column=1, value=year_i)
        p10_cell = sheet.cell(row=idx, column=2, value=float(result.p10[year_i]))
        p25_cell = sheet.cell(row=idx, column=3, value=float(result.p25[year_i]))
        p50_cell = sheet.cell(row=idx, column=4, value=float(result.p50[year_i]))
        p75_cell = sheet.cell(row=idx, column=5, value=float(result.p75[year_i]))
        p90_cell = sheet.cell(row=idx, column=6, value=float(result.p90[year_i]))

        for value_cell in (p10_cell, p25_cell, p50_cell, p75_cell, p90_cell):
            value_cell.number_format = "$#,##0.00"

        if year_i == inputs.years_to_retirement:
            marker_cell = sheet.cell(row=idx, column=7, value="RETIREMENT BEGINS")
            marker_cell.fill = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")
            marker_cell.font = Font(bold=True)

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 24
    sheet.freeze_panes = "A2"


def _render_chart_image(result: SimulationResult, retirement_year: int) -> Path:
    _require_dependencies()
    assert plt is not None and FuncFormatter is not None

    fig, axis = plt.subplots(figsize=(11, 5.8))
    years = result.years

    axis.fill_between(years, result.p10, result.p90, color=BRAND_NAVY, alpha=0.20, label="10th-90th percentile")
    axis.fill_between(years, result.p25, result.p75, color=BRAND_NAVY, alpha=0.35, label="25th-75th percentile")
    axis.plot(years, result.p50, color=BRAND_NAVY, linewidth=2.0, label="Median")
    axis.axvline(retirement_year, color="#3d5f8e", linestyle="--", linewidth=1.5)

    axis.set_title("Monte Carlo Simulation - Portfolio Projection")
    axis.set_xlabel("Years")
    axis.set_ylabel("Portfolio value ($)")
    axis.grid(alpha=0.18)
    axis.legend(loc="upper left")

    ylim = axis.get_ylim()
    axis.text(min(retirement_year + 0.6, years[-1]), ylim[1] * 0.95, "Retirement", color="#2c476d", fontsize=10)

    axis.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"${value:,.0f}"))
    fig.tight_layout()

    temp_file = tempfile.NamedTemporaryFile(prefix="mc_chart_", suffix=".png", delete=False)
    temp_path = Path(temp_file.name)
    temp_file.close()
    fig.savefig(temp_path, dpi=150)
    plt.close(fig)
    return temp_path


def _write_chart_sheet(sheet, result: SimulationResult, retirement_year: int) -> Path:
    assert XLImage is not None

    chart_path = _render_chart_image(result, retirement_year)
    img = XLImage(str(chart_path))
    img.width = 980
    img.height = 520
    sheet.add_image(img, "A1")
    return chart_path


def write_excel_report(inputs: SimulationInputs, result: SimulationResult) -> None:
    _require_dependencies()
    assert Workbook is not None and load_workbook is not None

    workbook_path = inputs.workbook_path
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()

    summary_sheet = _replace_or_create_sheet(workbook, "MC_Summary")
    percentiles_sheet = _replace_or_create_sheet(workbook, "MC_Percentiles")
    chart_sheet = _replace_or_create_sheet(workbook, "MC_Chart")

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 3:
        del workbook["Sheet"]

    _write_summary_sheet(summary_sheet, inputs, result)
    _write_percentiles_sheet(percentiles_sheet, inputs, result)
    chart_path = _write_chart_sheet(chart_sheet, result, inputs.years_to_retirement)

    try:
        try:
            workbook.save(workbook_path)
        except PermissionError as exc:
            raise PermissionError(
                "The workbook appears to be open. Close it in Excel and try again."
            ) from exc
    finally:
        chart_path.unlink(missing_ok=True)


def write_csv_export(csv_path: Path, result: SimulationResult) -> None:
    today = dt.date.today()
    with csv_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["Year", "Date", "P10", "P25", "P50", "P75", "P90"])
        for year in result.years:
            year_i = int(year)
            writer.writerow(
                [
                    year_i,
                    _add_years_safe(today, year_i).isoformat(),
                    f"{float(result.p10[year_i]):.2f}",
                    f"{float(result.p25[year_i]):.2f}",
                    f"{float(result.p50[year_i]):.2f}",
                    f"{float(result.p75[year_i]):.2f}",
                    f"{float(result.p90[year_i]):.2f}",
                ]
            )


class MonteCarloApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry(WINDOW_SIZE)
        self.root.resizable(False, False)

        self.entries: dict[str, ttk.Entry] = {}
        self.workbook_path_var = tk.StringVar(value="")
        self.export_csv_var = tk.BooleanVar(value=False)
        self.csv_path_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="Ready. Fill in inputs and select a workbook.")

        self.worker_thread: threading.Thread | None = None
        self.worker_events: queue.Queue[tuple[str, object]] = queue.Queue()
        self.is_running = False

        self._build_ui()

    def _build_ui(self) -> None:
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Primary.TButton", foreground="#ffffff", background=STATUS_NEUTRAL)

        frame = ttk.Frame(self.root, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)

        fields = [
            ("Current Portfolio Value ($)", "current_portfolio", ""),
            ("Annual Contribution ($)", "annual_contribution", ""),
            ("Contribution Growth Rate (% / yr)", "contribution_growth_rate", "0"),
            ("Years to Retirement", "years_to_retirement", ""),
            ("Years in Retirement", "years_in_retirement", "25"),
            ("Expected Annual Return (%)", "expected_return", "7.0"),
            ("Annual Volatility / Std Dev (%)", "volatility", "12.0"),
            ("Inflation Rate (%)", "inflation_rate", "2.5"),
            ("Annual Retirement Spending ($)", "annual_spending", ""),
            ("CPP / OAS / Pension Income ($ / yr)", "pension_income", "0"),
            ("Number of Simulations", "simulations", "1000"),
        ]

        row = 0
        for label_text, key, default in fields:
            ttk.Label(frame, text=label_text).grid(row=row, column=0, sticky="w", pady=2, padx=(0, 8))
            entry = ttk.Entry(frame)
            entry.insert(0, default)
            entry.grid(row=row, column=1, columnspan=2, sticky="ew", pady=2)
            self.entries[key] = entry
            row += 1

        csv_check = ttk.Checkbutton(
            frame,
            text="Also export to CSV",
            variable=self.export_csv_var,
            command=self._toggle_csv_controls,
        )
        csv_check.grid(row=row, column=0, columnspan=3, sticky="w", pady=(6, 2))
        row += 1

        ttk.Label(frame, text="CSV Output Path").grid(row=row, column=0, sticky="w", pady=2, padx=(0, 8))
        self.csv_entry = ttk.Entry(frame, textvariable=self.csv_path_var, state="disabled")
        self.csv_entry.grid(row=row, column=1, sticky="ew", pady=2)
        self.csv_button = ttk.Button(frame, text="Browse...", command=self._browse_csv, state="disabled")
        self.csv_button.grid(row=row, column=2, sticky="ew", pady=2, padx=(6, 0))
        row += 1

        ttk.Label(frame, text="Target .xlsx Workbook").grid(row=row, column=0, sticky="w", pady=2, padx=(0, 8))
        self.workbook_entry = ttk.Entry(frame, textvariable=self.workbook_path_var)
        self.workbook_entry.grid(row=row, column=1, sticky="ew", pady=2)
        self.workbook_button = ttk.Button(frame, text="Browse...", command=self._browse_workbook)
        self.workbook_button.grid(row=row, column=2, sticky="ew", pady=2, padx=(6, 0))
        row += 1

        self.progress = ttk.Progressbar(frame, mode="indeterminate")
        self.progress.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(8, 6))
        row += 1

        self.run_button = ttk.Button(
            frame,
            text="Run Simulation",
            command=self._on_run_clicked,
            style="Primary.TButton",
        )
        self.run_button.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(2, 6))
        row += 1

        self.status_label = tk.Label(
            frame,
            textvariable=self.status_var,
            anchor="w",
            fg=STATUS_NEUTRAL,
            bg=self.root.cget("bg"),
        )
        self.status_label.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(2, 0))

    def _toggle_csv_controls(self) -> None:
        enabled = bool(self.export_csv_var.get())
        state = "normal" if enabled else "disabled"
        self.csv_entry.config(state=state)
        self.csv_button.config(state=state)
        if not enabled:
            self.csv_path_var.set("")

    def _browse_workbook(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Select target workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if path:
            self.workbook_path_var.set(path)

    def _browse_csv(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Select CSV output",
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv")],
        )
        if path:
            self.csv_path_var.set(path)

    def _set_status(self, message: str, color: str = STATUS_NEUTRAL) -> None:
        self.status_var.set(message)
        self.status_label.config(fg=color)

    def _show_validation_error(self, message: str) -> None:
        self._set_status(f"Error: {message}", STATUS_ERROR)
        messagebox.showerror(WINDOW_TITLE, message)

    def _parse_float(self, key: str, message: str, allow_empty: bool = False) -> float:
        raw = self.entries[key].get().strip()
        if raw == "":
            if allow_empty:
                return 0.0
            raise ValidationError(message)
        try:
            return float(raw)
        except ValueError as exc:
            raise ValidationError(message) from exc

    def _parse_int(self, key: str, message: str) -> int:
        raw = self.entries[key].get().strip()
        if raw == "":
            raise ValidationError(message)
        try:
            value = int(raw)
        except ValueError as exc:
            raise ValidationError(message) from exc
        return value

    def _validate_inputs(self) -> SimulationInputs:
        current_portfolio = self._parse_float(
            "current_portfolio",
            "Current portfolio value must be greater than 0.",
        )
        if current_portfolio <= 0:
            raise ValidationError("Current portfolio value must be greater than 0.")

        annual_contribution = self._parse_float(
            "annual_contribution",
            "Enter annual contribution (0 is valid).",
        )

        contribution_growth_rate = self._parse_float(
            "contribution_growth_rate",
            "Contribution growth rate must be a number (0 is valid).",
        )

        years_to_retirement = self._parse_int(
            "years_to_retirement",
            "Years to retirement must be a whole number of 1 or more.",
        )
        if years_to_retirement < 1:
            raise ValidationError("Years to retirement must be a whole number of 1 or more.")

        years_in_retirement = self._parse_int(
            "years_in_retirement",
            "Years in retirement must be a whole number of 1 or more.",
        )
        if years_in_retirement < 1:
            raise ValidationError("Years in retirement must be a whole number of 1 or more.")

        expected_return = self._parse_float(
            "expected_return",
            "Expected annual return must be a number.",
        )

        volatility = self._parse_float(
            "volatility",
            "Volatility must be 0 or greater.",
        )
        if volatility < 0:
            raise ValidationError("Volatility must be 0 or greater.")

        inflation_rate = self._parse_float(
            "inflation_rate",
            "Inflation rate must be a number (0 is valid).",
        )

        annual_spending = self._parse_float(
            "annual_spending",
            "Annual retirement spending must be greater than 0.",
        )
        if annual_spending <= 0:
            raise ValidationError("Annual retirement spending must be greater than 0.")

        pension_income = self._parse_float(
            "pension_income",
            "CPP/OAS/Pension income must be a number (0 is valid).",
        )

        simulations = self._parse_int(
            "simulations",
            "Number of simulations must be a whole number between 1 and 10,000.",
        )
        if simulations < 1 or simulations > 10000:
            raise ValidationError("Number of simulations must be a whole number between 1 and 10,000.")

        workbook_raw = self.workbook_path_var.get().strip()
        if not workbook_raw.lower().endswith(".xlsx"):
            raise ValidationError("Please select a target .xlsx workbook before running.")
        workbook_path = Path(workbook_raw)

        csv_path: Path | None = None
        if self.export_csv_var.get():
            csv_raw = self.csv_path_var.get().strip()
            if not csv_raw:
                raise ValidationError("Please select a CSV output path.")
            csv_path = Path(csv_raw)

        return SimulationInputs(
            current_portfolio=current_portfolio,
            annual_contribution=annual_contribution,
            contribution_growth_rate=contribution_growth_rate,
            years_to_retirement=years_to_retirement,
            years_in_retirement=years_in_retirement,
            expected_return=expected_return,
            volatility=volatility,
            inflation_rate=inflation_rate,
            annual_spending=annual_spending,
            pension_income=pension_income,
            simulations=simulations,
            workbook_path=workbook_path,
            csv_path=csv_path,
        )

    def _set_running_state(self, running: bool) -> None:
        self.is_running = running
        run_state = "disabled" if running else "normal"
        browse_state = "disabled" if running else "normal"
        self.run_button.config(state=run_state)
        self.workbook_button.config(state=browse_state)
        self.workbook_entry.config(state=("disabled" if running else "normal"))

        if running:
            self.progress.start(12)
        else:
            self.progress.stop()
            self.workbook_entry.config(state="normal")
        self._toggle_csv_controls()
        if running and self.export_csv_var.get():
            self.csv_entry.config(state="disabled")
            self.csv_button.config(state="disabled")

    def _on_run_clicked(self) -> None:
        if self.is_running:
            return

        try:
            _require_dependencies()
        except RuntimeError as err:
            self._show_validation_error(str(err))
            return

        try:
            inputs = self._validate_inputs()
        except ValidationError as err:
            self._show_validation_error(str(err))
            return

        self._set_running_state(True)
        self._set_status(
            f"Running {inputs.simulations:,} simulations... please wait.",
            STATUS_NEUTRAL,
        )

        self.worker_thread = threading.Thread(
            target=self._run_worker,
            args=(inputs,),
            daemon=True,
        )
        self.worker_thread.start()
        self.root.after(80, self._poll_worker_events)

    def _run_worker(self, inputs: SimulationInputs) -> None:
        try:
            result = run_monte_carlo(inputs)
            self.worker_events.put(("status", "Writing results to workbook..."))
            write_excel_report(inputs, result)
            if inputs.csv_path is not None:
                write_csv_export(inputs.csv_path, result)
            self.worker_events.put(("success", (inputs, result)))
        except PermissionError:
            self.worker_events.put(
                ("error", "The workbook appears to be open. Close it in Excel and try again.")
            )
        except Exception as err:  # pylint: disable=broad-except
            self.worker_events.put(("error", str(err)))

    def _poll_worker_events(self) -> None:
        while True:
            try:
                event, payload = self.worker_events.get_nowait()
            except queue.Empty:
                break

            if event == "status":
                self._set_status(str(payload), STATUS_NEUTRAL)
            elif event == "success":
                inputs, _result = payload  # type: ignore[misc]
                self._set_running_state(False)
                self._set_status(
                    f"Done. {inputs.simulations:,} simulations complete. Workbook updated.",
                    STATUS_SUCCESS,
                )
            elif event == "error":
                self._set_running_state(False)
                error_message = str(payload)
                self._set_status(f"Error: {error_message}", STATUS_ERROR)
                messagebox.showerror(WINDOW_TITLE, error_message)

        if self.is_running:
            self.root.after(80, self._poll_worker_events)


def main() -> None:
    if not TK_AVAILABLE:
        raise RuntimeError(
            "tkinter is not available in this Python environment. "
            "Install tkinter support to run the desktop GUI."
        )

    root = tk.Tk()
    MonteCarloApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
