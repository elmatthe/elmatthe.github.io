"""Export helpers for analysis outputs."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image


PLOT_EXPORT_FILENAMES = {
    "Indexed Price": "indexed_price_chart",
    "Cumulative Return": "cumulative_return_chart",
    "Correlation Heatmap": "correlation_heatmap",
    "Risk / Return Scatter": "risk_return_scatter",
    "Drawdown Chart": "drawdown_chart",
    "Rolling Volatility": "rolling_volatility",
    "Rolling Correlation": "rolling_correlation_vs_benchmark",
    "Regression Scatter": "regression_scatter",
}


def export_results_to_excel(
    output_path: str,
    metrics: pd.DataFrame,
    corr_matrix: pd.DataFrame,
    regression_results: pd.DataFrame,
    warnings: list[str],
    plot_paths: dict[str, str] | None = None,
) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    export_warnings = list(warnings)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="Dashboard Metrics", index=False)
        corr_matrix.to_excel(writer, sheet_name="Correlation Matrix")
        regression_results.to_excel(writer, sheet_name="Regression Results", index=False)
        workbook = writer.book
        charts_sheet = workbook.create_sheet("Charts")
        row = 1
        for chart_name, chart_path in (plot_paths or {}).items():
            charts_sheet.cell(row=row, column=1, value=chart_name)
            image_path = Path(chart_path)
            if not image_path.exists():
                note = f"Chart skipped because file was missing: {chart_name} ({image_path})"
                charts_sheet.cell(row=row + 1, column=1, value=note)
                export_warnings.append(note)
                row += 4
                continue
            try:
                image = ExcelImage(str(image_path))
                image.width = 640
                image.height = 390
                charts_sheet.add_image(image, f"A{row + 1}")
                row += 24
            except Exception as exc:
                note = f"Chart skipped because Excel image insertion failed: {chart_name} ({exc})"
                charts_sheet.cell(row=row + 1, column=1, value=note)
                export_warnings.append(note)
                row += 4
        pd.DataFrame({"Warning": export_warnings}).to_excel(writer, sheet_name="Warnings", index=False)
    return str(path)


def export_results_to_csv_folder(
    output_folder: str,
    metrics: pd.DataFrame,
    corr_matrix: pd.DataFrame,
    regression_results: pd.DataFrame,
    warnings: list[str],
) -> str:
    folder = Path(output_folder)
    folder.mkdir(parents=True, exist_ok=True)
    metrics.to_csv(folder / "dashboard_metrics.csv", index=False)
    corr_matrix.to_csv(folder / "correlation_matrix.csv")
    regression_results.to_csv(folder / "regression_results.csv", index=False)
    pd.DataFrame({"Warning": warnings}).to_csv(folder / "warnings.csv", index=False)
    return str(folder)


def export_plot_images(
    output_folder: str | Path,
    plot_paths: dict[str, str],
    timestamp: str,
    export_jpg: bool = True,
    export_png: bool = False,
) -> Path:
    folder = Path(output_folder) / f"stock_analysis_images_{timestamp}"
    folder.mkdir(parents=True, exist_ok=True)
    notes = []
    for chart_name, chart_path in (plot_paths or {}).items():
        source = Path(chart_path)
        base_name = PLOT_EXPORT_FILENAMES.get(chart_name, chart_name.lower().replace(" ", "_").replace("/", "_"))
        if not source.exists():
            notes.append(f"Skipped missing image: {chart_name} ({source})")
            continue
        if export_png:
            try:
                (folder / f"{base_name}.png").write_bytes(source.read_bytes())
            except Exception as exc:
                notes.append(f"Could not copy PNG for {chart_name}: {exc}")
        if export_jpg:
            try:
                with Image.open(source) as image:
                    converted = image.convert("RGBA")
                    background = Image.new("RGBA", converted.size, (255, 255, 255, 255))
                    background.alpha_composite(converted)
                    background.convert("RGB").save(folder / f"{base_name}.jpg", "JPEG", quality=92)
            except Exception as exc:
                notes.append(f"Could not export JPG for {chart_name}: {exc}")
    if notes:
        (folder / "image_export_warnings.txt").write_text("\n".join(notes), encoding="utf-8")
    return folder
