from pathlib import Path
import shutil
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from exports import export_plot_images, export_results_to_csv_folder, export_results_to_excel  # noqa: E402
from plots import plot_risk_return_scatter  # noqa: E402


def sample_outputs():
    metrics = pd.DataFrame(
        {
            "Ticker": ["AAPL"],
            "Total Return": [0.05],
            "Annualized Return": [0.10],
            "Annualized Volatility": [0.20],
            "Sharpe Ratio": [0.5],
            "Max Drawdown": [-0.03],
            "Observations": [10],
            "Data Completeness": ["Complete"],
            "Currency": ["USD"],
        }
    )
    corr = pd.DataFrame({"AAPL": [1.0]}, index=["AAPL"])
    regressions = pd.DataFrame({"Y Ticker": ["AAPL"], "X Ticker": ["SPY"], "beta": [1.1]})
    return metrics, corr, regressions, ["sample warning"]


def test_excel_export_creates_workbook():
    metrics, corr, regressions, warnings = sample_outputs()
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-excel"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    output = export_results_to_excel(str(temp_dir / "results.xlsx"), metrics, corr, regressions, warnings)
    assert Path(output).exists()
    workbook = load_workbook(output)
    assert {"Dashboard Metrics", "Correlation Matrix", "Regression Results", "Warnings", "Charts"}.issubset(workbook.sheetnames)


def test_excel_export_inserts_chart_sheet_images():
    metrics, corr, regressions, warnings = sample_outputs()
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-excel-chart"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    chart_path = plot_risk_return_scatter(metrics, str(temp_dir / "risk_return.png"))
    output = export_results_to_excel(
        str(temp_dir / "results.xlsx"),
        metrics,
        corr,
        regressions,
        warnings,
        {"Risk / Return Scatter": chart_path},
    )
    workbook = load_workbook(output)
    assert "Charts" in workbook.sheetnames
    assert len(workbook["Charts"]._images) == 1


def test_excel_export_skips_missing_plot_paths():
    metrics, corr, regressions, warnings = sample_outputs()
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-excel-missing-chart"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    output = export_results_to_excel(
        str(temp_dir / "results.xlsx"),
        metrics,
        corr,
        regressions,
        warnings,
        {"Missing Chart": str(temp_dir / "missing.png")},
    )
    workbook = load_workbook(output)
    warning_values = [cell.value for cell in workbook["Warnings"]["A"] if cell.value]
    assert any("Chart skipped" in value for value in warning_values)


def test_csv_export_creates_expected_files():
    metrics, corr, regressions, warnings = sample_outputs()
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-csv"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    folder = Path(export_results_to_csv_folder(str(temp_dir), metrics, corr, regressions, warnings))
    assert (folder / "dashboard_metrics.csv").exists()
    assert (folder / "correlation_matrix.csv").exists()
    assert (folder / "regression_results.csv").exists()
    assert (folder / "warnings.csv").exists()


def test_plot_image_export_creates_jpg_folder():
    metrics, _corr, _regressions, _warnings = sample_outputs()
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-images"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    chart_path = plot_risk_return_scatter(metrics, str(temp_dir / "risk_return.png"))
    folder = export_plot_images(temp_dir, {"Risk / Return Scatter": chart_path}, "20260609_120000")
    assert (folder / "risk_return_scatter.jpg").exists()


def test_plot_image_export_skips_missing_paths_without_crashing():
    temp_dir = Path(__file__).resolve().parents[1] / ".run-temp" / "exports-images-missing"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True)
    folder = export_plot_images(temp_dir, {"Missing Chart": str(temp_dir / "missing.png")}, "20260609_120000")
    assert folder.exists()
    assert (folder / "image_export_warnings.txt").exists()
